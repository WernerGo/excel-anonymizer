"""Shared fixtures for all tests."""

import pytest
import openpyxl
import yaml
from pathlib import Path


@pytest.fixture
def sample_excel(tmp_path: Path) -> Path:
    """
    Excel with employee/manager columns where the same names appear in both.

    Sheet1 (data from row 2):
      col A: employee last name   col B: employee first name
      col C: manager last name    col D: manager first name

    Row 2: Smith, John  → managed by Brown, Jane
    Row 3: Brown, Jane  → managed by Smith, John

    "Smith" appears in A2 and C3 → must get the same key (relationship test).
    "Brown" appears in A3 and C2 → must get the same key.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(1, 1, "Last Name");  ws.cell(1, 2, "First Name")
    ws.cell(1, 3, "Mgr Last");   ws.cell(1, 4, "Mgr First")
    ws.cell(2, 1, "Smith");  ws.cell(2, 2, "John")
    ws.cell(2, 3, "Brown");  ws.cell(2, 4, "Jane")
    ws.cell(3, 1, "Brown");  ws.cell(3, 2, "Jane")
    ws.cell(3, 3, "Smith");  ws.cell(3, 4, "John")
    path = tmp_path / "employees.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def anon_config(tmp_path: Path) -> Path:
    """Config for anonymize.py with two independent groups."""
    config = {
        "output_suffix": "_anonymized",
        "save_mapping": "anon_map.json",
        "groups": [
            {
                "name": "last_names",
                "prefix": "LAST",
                "columns": [
                    {"sheet": "Sheet1", "col": "A", "data_from_row": 2},
                    {"sheet": "Sheet1", "col": "C", "data_from_row": 2},
                ],
            },
            {
                "name": "first_names",
                "prefix": "FIRST",
                "columns": [
                    {"sheet": "Sheet1", "col": "B", "data_from_row": 2},
                    {"sheet": "Sheet1", "col": "D", "data_from_row": 2},
                ],
            },
        ],
    }
    path = tmp_path / "config.yaml"
    path.write_text(yaml.dump(config), encoding="utf-8")
    return path


@pytest.fixture
def faker_config(tmp_path: Path) -> Path:
    """Config for faker_replace.py."""
    config = {
        "output_suffix": "_faker",
        "save_mapping": "faker_map.json",
        "locale": "en_US",
        "groups": [
            {
                "name": "last_names",
                "faker_type": "last_name",
                "columns": [
                    {"sheet": "Sheet1", "col": "A", "data_from_row": 2},
                    {"sheet": "Sheet1", "col": "C", "data_from_row": 2},
                ],
            },
            {
                "name": "first_names",
                "faker_type": "first_name",
                "columns": [
                    {"sheet": "Sheet1", "col": "B", "data_from_row": 2},
                    {"sheet": "Sheet1", "col": "D", "data_from_row": 2},
                ],
            },
        ],
    }
    path = tmp_path / "faker_config.yaml"
    path.write_text(yaml.dump(config), encoding="utf-8")
    return path
