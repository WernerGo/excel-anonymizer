"""Tests for faker_replace.py: realistic fake data replacement."""

import json
import openpyxl
import yaml
import pytest
from pathlib import Path

from faker_replace import faker_replace


def _load_output(excel_path: Path, suffix: str) -> openpyxl.worksheet.worksheet.Worksheet:
    out = excel_path.with_stem(excel_path.stem + suffix)
    return openpyxl.load_workbook(out)["Sheet1"]


# ---------------------------------------------------------------------------
# Basic replacement
# ---------------------------------------------------------------------------

def test_original_values_are_gone(sample_excel, faker_config):
    faker_replace(sample_excel, faker_config)
    ws = _load_output(sample_excel, "_faker")
    data_values = [ws.cell(r, c).value for r in (2, 3) for c in (1, 2, 3, 4)]
    assert "Smith" not in data_values
    assert "Brown" not in data_values
    assert "John"  not in data_values
    assert "Jane"  not in data_values


# ---------------------------------------------------------------------------
# Relationship preservation
# ---------------------------------------------------------------------------

def test_same_value_gets_same_fake_across_columns(sample_excel, faker_config):
    """Smith in col A and col C must get the same fake last name."""
    faker_replace(sample_excel, faker_config)
    ws = _load_output(sample_excel, "_faker")
    assert ws.cell(2, 1).value == ws.cell(3, 3).value  # Smith → same fake
    assert ws.cell(3, 1).value == ws.cell(2, 3).value  # Brown → same fake


def test_different_values_get_different_fakes(sample_excel, faker_config):
    faker_replace(sample_excel, faker_config)
    ws = _load_output(sample_excel, "_faker")
    assert ws.cell(2, 1).value != ws.cell(3, 1).value  # Smith ≠ Brown


# ---------------------------------------------------------------------------
# Uniqueness guarantee (core correctness requirement)
# ---------------------------------------------------------------------------

def test_no_two_originals_share_the_same_fake(tmp_path):
    """All unique originals must map to distinct fake values — no collisions."""
    originals = [f"Person{i:03d}" for i in range(50)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(1, 1, "Name")
    for i, name in enumerate(originals, start=2):
        ws.cell(i, 1, name)
    excel_path = tmp_path / "many_names.xlsx"
    wb.save(excel_path)

    config = {
        "output_suffix": "_faker",
        "save_mapping": "faker_map.json",
        "locale": "en_US",
        "groups": [{"name": "names", "faker_type": "last_name",
                    "columns": [{"sheet": "Sheet1", "col": "A", "data_from_row": 2}]}],
    }
    config_path = tmp_path / "config.yaml"
    config_path.write_text(yaml.dump(config), encoding="utf-8")

    faker_replace(excel_path, config_path)

    mapping = json.loads((tmp_path / "faker_map.json").read_text())
    fake_values = list(mapping["names"].values())
    assert len(fake_values) == len(set(fake_values)), "Collision detected: two originals got the same fake value"


def test_pool_exhaustion_still_produces_unique_fakes(tmp_path):
    """When more unique values than Faker's pool, fallback suffix keeps uniqueness."""
    # 300 unique values exceeds any Faker locale's last name pool
    originals = [f"UniqueLastName{i:04d}" for i in range(300)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(1, 1, "Name")
    for i, name in enumerate(originals, start=2):
        ws.cell(i, 1, name)
    excel_path = tmp_path / "large.xlsx"
    wb.save(excel_path)

    config = {
        "output_suffix": "_faker",
        "save_mapping": "faker_map.json",
        "locale": "en_US",
        "groups": [{"name": "names", "faker_type": "last_name",
                    "columns": [{"sheet": "Sheet1", "col": "A", "data_from_row": 2}]}],
    }
    config_path = tmp_path / "config.yaml"
    config_path.write_text(yaml.dump(config), encoding="utf-8")

    faker_replace(excel_path, config_path)

    mapping = json.loads((tmp_path / "faker_map.json").read_text())
    fake_values = list(mapping["names"].values())
    assert len(fake_values) == len(set(fake_values)), "Collision after pool exhaustion"


# ---------------------------------------------------------------------------
# Roundtrip: faker_replace → deanonymize
# ---------------------------------------------------------------------------

def test_roundtrip_restores_original_values(sample_excel, faker_config):
    from deanonymize import deanonymize

    faker_replace(sample_excel, faker_config)
    faker_path = sample_excel.with_stem(sample_excel.stem + "_faker")
    map_path   = sample_excel.parent / "faker_map.json"

    deanonymize(faker_path, map_path)
    restored_path = faker_path.with_stem(faker_path.stem + "_restored")

    original = openpyxl.load_workbook(sample_excel)["Sheet1"]
    restored = openpyxl.load_workbook(restored_path)["Sheet1"]

    for row in range(1, 4):
        for col in range(1, 5):
            assert original.cell(row, col).value == restored.cell(row, col).value
