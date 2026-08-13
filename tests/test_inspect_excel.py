"""What the inspector has to report before a mapping can be written.

The two questions it answers are not the same one. For a migration a
formula cell is derived and is not taken over. For an anonymisation it
does not matter that a formula stood there: resolving it leaves the real
figure in the cell, and that has to be replaced like any other.
"""

import zipfile
from pathlib import Path

import openpyxl
import pytest

from inspect_excel import column_analysis, parse_header_rows, workbook_info


@pytest.fixture
def excel_with_a_title_row(tmp_path: Path) -> Path:
    """A workbook laid out like the PE Tracker: a title, then headings.

    openpyxl cannot write a calculated result, so the formula cells here
    stand for a workbook Excel has never recalculated — which is the
    conservative case for the test.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Transactions"
    ws.append(["PE Tracker — Cash Transactions"])
    ws.append(["investment_id", "amount", "rating"])
    ws.append([1, 1000, "BBB"])
    ws.append([2, 2000, "=A3"])
    path = tmp_path / "book.xlsx"
    wb.save(path)
    return path


def test_the_header_row_can_be_named_per_sheet(excel_with_a_title_row):
    """Not every workbook puts its headings in the first row, and the
    ones this was written for do not put them in the same row twice."""
    manifest = workbook_info(
        excel_with_a_title_row,
        header_rows={"Cash Transactions": 2},
    )
    columns = manifest["worksheets"][0]["columns"]

    assert [c["header"] for c in columns.values()] == ["investment_id", "amount", "rating"]
    assert columns["A"]["value_count"] == 2, "the title and the headings are not data"


def test_without_a_header_row_the_headings_count_as_data(excel_with_a_title_row):
    """The default is unchanged: read everything, name nothing."""
    columns = workbook_info(excel_with_a_title_row)["worksheets"][0]["columns"]

    assert "header" not in columns["A"]
    assert columns["A"]["value_count"] == 4


def with_a_cached_result(path: Path, target: Path, formula: bytes, result: bytes) -> Path:
    """Give a formula cell the result Excel would have stored for it.

    openpyxl writes the formula and no result, so a file it produced has
    no cached values at all. Every real workbook has them, and they are
    what the inspector reports the types of.
    """
    with zipfile.ZipFile(path) as reader, zipfile.ZipFile(target, "w") as writer:
        for item in reader.infolist():
            data = reader.read(item.filename)
            if item.filename.startswith("xl/worksheets/sheet"):
                data = data.replace(formula, formula[:-len(b"</f>")] + b"</f>" + result)
            writer.writestr(item, data)
    return target


def test_a_formula_is_counted_and_its_result_is_typed(tmp_path):
    """A column part formula and part typed in is where data hides.

    445 columns of one sample workbook are of that kind, so "the column
    holds formulas, therefore it is derived" would drop original values
    in hundreds of places.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "positions"
    ws.append(["transfer_lgd"])
    ws.append([0.45])
    ws.append(["=1-0.6"])
    written = tmp_path / "written.xlsx"
    wb.save(written)

    path = with_a_cached_result(
        written, tmp_path / "book.xlsx", b"<f>1-0.6</f>", b"<v>0.4</v>"
    )
    column = workbook_info(path, header_rows={"*": 1})["worksheets"][0]["columns"]["A"]

    assert column["formula_count"] == 1
    assert column["value_count"] == 2, "one typed in, one calculated"
    assert column["formula_count"] != column["value_count"], "so the column is mixed"
    assert column["types"] == {"number": 2}, "the type of the result, not 'formula'"


def test_a_formula_that_was_never_calculated_holds_no_value(tmp_path):
    """It counts as a formula and not as a value, which is the truth:
    in a file Excel has not recalculated the cell is empty."""
    wb = openpyxl.Workbook()
    wb.active.append(["transfer_lgd"])
    wb.active.append(["=1-0.6"])
    path = tmp_path / "book.xlsx"
    wb.save(path)

    column = workbook_info(path, header_rows={"*": 1})["worksheets"][0]["columns"]["A"]

    assert column["formula_count"] == 1
    assert column["value_count"] == 0


def test_a_column_without_formulas_says_nothing_about_them(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.append(["amount"])
    wb.active.append([1000])
    path = tmp_path / "book.xlsx"
    wb.save(path)

    assert "formula_count" not in workbook_info(path)["worksheets"][0]["columns"]["A"]


def test_the_cell_by_cell_dump_is_off_by_default(excel_with_a_title_row):
    """One sheet of the PE Tracker holds 108,532 rows."""
    sheet = workbook_info(excel_with_a_title_row)["worksheets"][0]

    assert "cells" not in sheet
    assert "cells" in workbook_info(excel_with_a_title_row, include_cells=True)["worksheets"][0]


def test_an_xlsm_is_read_like_an_xlsx(excel_with_a_title_row, tmp_path):
    """.xlsm is .xlsx with macros beside it. The data is what is wanted."""
    macro_file = tmp_path / "book.xlsm"
    macro_file.write_bytes(excel_with_a_title_row.read_bytes())

    manifest = workbook_info(macro_file)

    assert manifest["format"] == "xlsm"
    assert manifest["worksheets"][0]["name"] == "Cash Transactions"


def test_header_rows_are_read_for_one_sheet_or_for_all():
    assert parse_header_rows(["3"]) == {"*": 3}
    assert parse_header_rows(["Cash Transactions=3", "FA_Support=6"]) == {
        "Cash Transactions": 3,
        "FA_Support": 6,
    }
    assert parse_header_rows(None) == {}


def test_a_header_row_that_is_not_a_number_stops_the_run():
    with pytest.raises(SystemExit, match="SHEET=N"):
        parse_header_rows(["Cash Transactions=three"])
