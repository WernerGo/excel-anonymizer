"""Tests for the two settings that decide what survives a run.

``formulas`` decides whether the output holds formulas or their results,
and one of the two is always lost. ``strategy: scale`` makes amounts
fictitious without destroying the relations between them.
"""

import datetime
import zipfile

import openpyxl
import pytest
import yaml
from pathlib import Path

from anonymize import anonymize, count_uncached_formulas, load_source, scale_value


@pytest.fixture
def excel_with_formulas(tmp_path: Path) -> Path:
    """A workbook whose amounts are partly typed in and partly calculated.

    openpyxl cannot write a calculated result, only the formula, so these
    cells stand for the case of a file that Excel has never recalculated.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(1, 1, "name");   ws.cell(1, 2, "amount");  ws.cell(1, 3, "total")
    ws.cell(2, 1, "Smith");  ws.cell(2, 2, 1000);      ws.cell(2, 3, "=B2*2")
    ws.cell(3, 1, "Brown");  ws.cell(3, 2, 1500.5);    ws.cell(3, 3, "=B3*2")
    path = tmp_path / "amounts.xlsx"
    wb.save(path)
    return path


def write_config(tmp_path: Path, config: dict) -> Path:
    """Write a config and return its path."""
    path = tmp_path / "config.yaml"
    path.write_text(yaml.dump(config), encoding="utf-8")
    return path


# ---------------------------------------------------------------------------
# formulas: values | keep
# ---------------------------------------------------------------------------

def test_values_is_the_default(excel_with_formulas, tmp_path):
    """Most users want data out of the file, so resolving is the default."""
    config = write_config(tmp_path, {"groups": []})
    anonymize(excel_with_formulas, config)

    out = excel_with_formulas.with_stem(excel_with_formulas.stem + "_anonymized")
    ws = openpyxl.load_workbook(out)["Sheet1"]
    assert ws.cell(2, 3).value != "=B2*2", "the formula must not survive in values mode"


def test_keep_leaves_the_formulas_in_place(excel_with_formulas, tmp_path):
    """For a workbook that stays in use in Excel."""
    config = write_config(tmp_path, {"formulas": "keep", "groups": []})
    anonymize(excel_with_formulas, config)

    out = excel_with_formulas.with_stem(excel_with_formulas.stem + "_anonymized")
    ws = openpyxl.load_workbook(out)["Sheet1"]
    assert ws.cell(2, 3).value == "=B2*2"


def test_an_unknown_mode_is_refused(excel_with_formulas):
    """A typo here silently changes what the output contains."""
    with pytest.raises(ValueError, match="formulas"):
        load_source(excel_with_formulas, "cached")


def rewrite_sheets(source: Path, target: Path, change) -> Path:
    """Copy a workbook, passing every worksheet part through `change`."""
    with zipfile.ZipFile(source) as reader, zipfile.ZipFile(target, "w") as writer:
        for item in reader.infolist():
            data = reader.read(item.filename)
            if item.filename.startswith("xl/worksheets/sheet"):
                data = change(data)
            writer.writestr(item, data)
    return target


def test_a_formula_that_was_never_calculated_is_counted(excel_with_formulas, tmp_path):
    """A workbook Excel has never calculated stores the formula and no result.

    That is the case worth warning about: in `values` mode those cells
    arrive empty and nothing says they ever held anything.
    """
    never = rewrite_sheets(
        excel_with_formulas, tmp_path / "never.xlsx", lambda data: data.replace(b"<v />", b"")
    )

    assert count_uncached_formulas(never) == 2


def test_a_formula_whose_result_is_empty_is_not_counted(excel_with_formulas):
    """`=IFERROR(…, "")` is calculated and its answer is nothing.

    Excel records that as an empty value element and openpyxl reports it
    as None — the same as never calculated. Counting the two together
    made a well-formed workbook look broken: one real file reported 1061
    such cells, every one of them correct. openpyxl writes its own
    formula cells the same way, which is why this fixture stands for the
    case.
    """
    assert count_uncached_formulas(excel_with_formulas) == 0


def test_a_file_without_formulas_counts_none(tmp_path):
    """No warning where there is nothing to warn about."""
    wb = openpyxl.Workbook()
    wb.active.cell(1, 1, "plain")
    path = tmp_path / "plain.xlsx"
    wb.save(path)

    assert count_uncached_formulas(path) == 0


# ---------------------------------------------------------------------------
# strategy: scale
# ---------------------------------------------------------------------------

def test_an_integer_stays_an_integer():
    """Writing 1283.7 where the source held 1000 changes the column's shape."""
    assert scale_value(1000, 1.2837) == 1284
    assert isinstance(scale_value(1000, 1.2837), int)


def test_a_float_stays_a_float():
    assert scale_value(1500.5, 2) == 3001.0
    assert isinstance(scale_value(1500.5, 2), float)


def test_a_flag_is_not_a_number():
    """`active_flag` is 0 or 1 in these files, and scaling it is nonsense."""
    assert scale_value(True, 2) is None


def test_text_and_dates_are_left_alone():
    """A numeric column holding `#N/A` is a finding, not something to scale."""
    import datetime

    assert scale_value("#N/A", 2) is None
    assert scale_value(datetime.datetime(2026, 3, 31), 2) is None


def test_scaling_keeps_the_relation_between_two_amounts(excel_with_formulas, tmp_path):
    """The point of one factor per workbook: sums and ratios still work out."""
    config = write_config(
        tmp_path,
        {
            "groups": [
                {
                    "name": "amounts",
                    "strategy": "scale",
                    "factor": 2,
                    "columns": [{"sheet": "Sheet1", "col": "B", "data_from_row": 2}],
                }
            ]
        },
    )
    anonymize(excel_with_formulas, config)

    out = excel_with_formulas.with_stem(excel_with_formulas.stem + "_anonymized")
    ws = openpyxl.load_workbook(out)["Sheet1"]
    assert ws.cell(2, 2).value == 2000
    assert ws.cell(3, 2).value == 3001.0
    assert ws.cell(3, 2).value / ws.cell(2, 2).value == pytest.approx(1500.5 / 1000)


def test_a_scale_group_needs_a_factor(excel_with_formulas, tmp_path):
    """Without one the group would silently do nothing."""
    config = write_config(
        tmp_path,
        {"groups": [{"name": "amounts", "strategy": "scale", "columns": []}]},
    )
    with pytest.raises(ValueError, match="factor"):
        anonymize(excel_with_formulas, config)


def test_a_key_group_needs_a_prefix(excel_with_formulas, tmp_path):
    """Same reasoning from the other side."""
    config = write_config(
        tmp_path,
        {"groups": [{"name": "names", "columns": []}]},
    )
    with pytest.raises(ValueError, match="prefix"):
        anonymize(excel_with_formulas, config)


def test_an_unknown_strategy_is_refused(excel_with_formulas, tmp_path):
    """A misspelt strategy would leave the column untouched and unreported."""
    config = write_config(
        tmp_path,
        {"groups": [{"name": "amounts", "strategy": "scaled", "columns": []}]},
    )
    with pytest.raises(ValueError, match="strategy"):
        anonymize(excel_with_formulas, config)


def test_scale_groups_are_not_written_to_the_mapping(excel_with_formulas, tmp_path):
    """Rounding loses the remainder, so the mapping could not restore them anyway."""
    import json

    config = write_config(
        tmp_path,
        {
            "save_mapping": "map.json",
            "groups": [
                {
                    "name": "names",
                    "prefix": "NAME",
                    "columns": [{"sheet": "Sheet1", "col": "A", "data_from_row": 2}],
                },
                {
                    "name": "amounts",
                    "strategy": "scale",
                    "factor": 2,
                    "columns": [{"sheet": "Sheet1", "col": "B", "data_from_row": 2}],
                },
            ],
        },
    )
    anonymize(excel_with_formulas, config)

    mapping = json.loads((excel_with_formulas.parent / "map.json").read_text())
    assert set(mapping) == {"names"}


# ---------------------------------------------------------------------------
# ignore_sheets
# ---------------------------------------------------------------------------

@pytest.fixture
def excel_with_a_working_sheet(tmp_path: Path) -> Path:
    """A workbook with a data sheet and a working view repeating its values."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.cell(1, 1, "name");  ws.cell(2, 1, "Smith")

    view = wb.create_sheet("Overview")
    view.cell(1, 1, "name");  view.cell(2, 1, "Smith")

    path = tmp_path / "book.xlsx"
    wb.save(path)
    return path


def test_a_listed_sheet_is_removed(excel_with_a_working_sheet, tmp_path):
    """The whole point: a sheet that is gone cannot carry an original value."""
    config = write_config(
        tmp_path,
        {"ignore_sheets": [{"sheet": "Overview", "reason": "a view of Data"}], "groups": []},
    )
    anonymize(excel_with_a_working_sheet, config)

    out = excel_with_a_working_sheet.with_stem(excel_with_a_working_sheet.stem + "_anonymized")
    assert openpyxl.load_workbook(out).sheetnames == ["Data"]


def test_a_plain_list_of_names_works_too(excel_with_a_working_sheet, tmp_path):
    """The reason is for the reader; the tool does not insist on it."""
    config = write_config(tmp_path, {"ignore_sheets": ["Overview"], "groups": []})
    anonymize(excel_with_a_working_sheet, config)

    out = excel_with_a_working_sheet.with_stem(excel_with_a_working_sheet.stem + "_anonymized")
    assert openpyxl.load_workbook(out).sheetnames == ["Data"]


def test_removing_every_sheet_is_refused(excel_with_a_working_sheet, tmp_path):
    """An empty workbook cannot be saved, and would not be wanted anyway."""
    config = write_config(tmp_path, {"ignore_sheets": ["Data", "Overview"], "groups": []})

    with pytest.raises(ValueError, match="every sheet"):
        anonymize(excel_with_a_working_sheet, config)


def test_a_sheet_that_is_not_there_only_warns(excel_with_a_working_sheet, tmp_path, capsys):
    """One list serves several files of the same family; not all carry every sheet."""
    config = write_config(tmp_path, {"ignore_sheets": ["Gone", "Overview"], "groups": []})
    anonymize(excel_with_a_working_sheet, config)

    assert "not in the file" in capsys.readouterr().out


def test_a_removed_sheet_takes_its_values_with_it(excel_with_a_working_sheet, tmp_path):
    """Anonymizing the data sheet alone would leave the copy in the view."""
    config = write_config(
        tmp_path,
        {
            "ignore_sheets": ["Overview"],
            "groups": [
                {"name": "names", "prefix": "NAME", "columns": [{"sheet": "Data", "col": "A", "data_from_row": 2}]}
            ],
        },
    )
    anonymize(excel_with_a_working_sheet, config)

    out = excel_with_a_working_sheet.with_stem(excel_with_a_working_sheet.stem + "_anonymized")
    book = openpyxl.load_workbook(out)
    values = [cell.value for sheet in book.worksheets for row in sheet.iter_rows() for cell in row]
    assert "Smith" not in values


# ---------------------------------------------------------------------------
# keys and numbers
# ---------------------------------------------------------------------------

def test_a_key_group_leaves_numbers_alone(excel_with_formulas, tmp_path):
    """A column of amounts with one text cell in it must stay a column of amounts."""
    config = write_config(
        tmp_path,
        {
            "groups": [
                {
                    "name": "mixed",
                    "prefix": "X",
                    "columns": [{"sheet": "Sheet1", "col": "B", "data_from_row": 2}],
                }
            ]
        },
    )
    anonymize(excel_with_formulas, config)

    ws = openpyxl.load_workbook(excel_with_formulas.with_stem(excel_with_formulas.stem + "_anonymized"))["Sheet1"]
    assert ws.cell(2, 2).value == 1000
    assert ws.cell(3, 2).value == 1500.5


def test_include_numbers_says_so_explicitly(excel_with_formulas, tmp_path):
    """A tax number is numeric and still has to go — but the config has to ask."""
    config = write_config(
        tmp_path,
        {
            "groups": [
                {
                    "name": "tax",
                    "prefix": "TAXNO",
                    "include_numbers": True,
                    "columns": [{"sheet": "Sheet1", "col": "B", "data_from_row": 2}],
                }
            ]
        },
    )
    anonymize(excel_with_formulas, config)

    ws = openpyxl.load_workbook(excel_with_formulas.with_stem(excel_with_formulas.stem + "_anonymized"))["Sheet1"]
    assert ws.cell(2, 2).value.startswith("TAXNO")


# ---------------------------------------------------------------------------
# what a removed sheet and a resolved formula leave behind
# ---------------------------------------------------------------------------

def test_a_defined_name_pointing_at_a_removed_sheet_goes_with_it(tmp_path):
    """Otherwise Excel asks about links on every open and cannot resolve them.

    Written in `keep` mode, because that is the only mode in which a
    defined name survives at all: where the formulas are resolved they
    all go, names being hand-written labels that say what a range is
    about. What this guards is the other thing — that removing a sheet
    does not leave a name pointing into nothing.
    """
    from openpyxl.workbook.defined_name import DefinedName

    wb = openpyxl.Workbook()
    wb.active.title = "Data"
    wb.active.cell(1, 1, "a")
    wb.create_sheet("Overview").cell(1, 1, "b")
    wb.defined_names["from_overview"] = DefinedName("from_overview", attr_text="Overview!$A$1")
    wb.defined_names["from_data"] = DefinedName("from_data", attr_text="Data!$A$1")
    path = tmp_path / "book.xlsx"
    wb.save(path)

    config = write_config(tmp_path, {"formulas": "keep", "ignore_sheets": ["Overview"], "groups": []})
    anonymize(path, config)

    out = openpyxl.load_workbook(path.with_stem(path.stem + "_anonymized"))
    assert "from_overview" not in out.defined_names
    assert "from_data" in out.defined_names, "a name for a kept sheet must survive"


def test_a_quoted_sheet_name_is_recognised_too(tmp_path):
    """A sheet name with a space is quoted in a reference."""
    from openpyxl.workbook.defined_name import DefinedName

    wb = openpyxl.Workbook()
    wb.active.title = "Data"
    wb.active.cell(1, 1, "a")
    wb.create_sheet("Auflistung PE").cell(1, 1, "b")
    wb.defined_names["listing"] = DefinedName("listing", attr_text="'Auflistung PE'!$A$1")
    path = tmp_path / "book.xlsx"
    wb.save(path)

    config = write_config(tmp_path, {"ignore_sheets": ["Auflistung PE"], "groups": []})
    anonymize(path, config)

    out = openpyxl.load_workbook(path.with_stem(path.stem + "_anonymized"))
    assert "listing" not in out.defined_names


def table_workbook(tmp_path: Path) -> Path:
    """A workbook with a table whose column carries a formula."""
    from openpyxl.worksheet.table import Table, TableColumn, TableFormula

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["amount", "double"])
    ws.append([1000, 2000])
    ws.append([2000, 4000])
    table = Table(
        displayName="Amounts",
        ref="A1:B3",
        tableColumns=[
            TableColumn(id=1, name="amount"),
            TableColumn(id=2, name="double", calculatedColumnFormula=TableFormula(attr_text="Amounts[[#This Row],[amount]]*2")),
        ],
    )
    ws.add_table(table)
    path = tmp_path / "table.xlsx"
    wb.save(path)
    return path


def test_a_resolved_column_leaves_no_column_formula_behind(tmp_path):
    """Excel flags every cell of the column as an inconsistent calculation."""
    path = table_workbook(tmp_path)
    config = write_config(tmp_path, {"groups": []})
    anonymize(path, config)

    out = openpyxl.load_workbook(path.with_stem(path.stem + "_anonymized"))
    formulas = [c.calculatedColumnFormula for t in out["Data"].tables.values() for c in t.tableColumns]
    assert formulas == [None, None]


def test_keeping_the_formulas_keeps_the_column_formula(tmp_path):
    """There the table and the cells still agree."""
    path = table_workbook(tmp_path)
    config = write_config(tmp_path, {"formulas": "keep", "groups": []})
    anonymize(path, config)

    out = openpyxl.load_workbook(path.with_stem(path.stem + "_anonymized"))
    formulas = [c.calculatedColumnFormula for t in out["Data"].tables.values() for c in t.tableColumns]
    assert any(formula is not None for formula in formulas)


# ---------------------------------------------------------------------------
# sheets nobody mentioned
# ---------------------------------------------------------------------------

def test_a_sheet_in_no_list_is_reported(excel_with_a_working_sheet, tmp_path, capsys):
    """The next export carries a sheet the config was never written against.

    The count goes to the screen and the names go to a file beside the
    workbook. A sheet is often named after the transaction it holds, so
    the list of the ones nobody has decided about is itself confidential
    — and it is exactly what someone reaches for when they want to ask
    about it.
    """
    config = write_config(
        tmp_path,
        {"groups": [{"name": "n", "prefix": "N", "columns": [{"sheet": "Data", "col": "A"}]}]},
    )
    anonymize(excel_with_a_working_sheet, config)

    printed = capsys.readouterr().out
    assert "1 sheets are in the file but in no group" in printed
    assert "Overview" not in printed

    listing = excel_with_a_working_sheet.with_name(f"{excel_with_a_working_sheet.stem}_unlisted_sheets.txt")
    assert listing.read_text(encoding="utf-8").split() == ["Overview"]


def test_fail_refuses_to_write_a_file_with_an_unlisted_sheet(excel_with_a_working_sheet, tmp_path):
    """Where the output must be free of originals, unconsidered is not good enough."""
    config = write_config(
        tmp_path,
        {
            "unlisted_sheets": "fail",
            "groups": [{"name": "n", "prefix": "N", "columns": [{"sheet": "Data", "col": "A"}]}],
        },
    )
    with pytest.raises(ValueError, match="unaccounted"):
        anonymize(excel_with_a_working_sheet, config)


def test_an_ignored_sheet_counts_as_accounted_for(excel_with_a_working_sheet, tmp_path):
    """Dropping a sheet is a decision about it, so it satisfies the check."""
    config = write_config(
        tmp_path,
        {
            "unlisted_sheets": "fail",
            "ignore_sheets": ["Overview"],
            "groups": [{"name": "n", "prefix": "N", "columns": [{"sheet": "Data", "col": "A"}]}],
        },
    )
    anonymize(excel_with_a_working_sheet, config)

    out = excel_with_a_working_sheet.with_stem(excel_with_a_working_sheet.stem + "_anonymized")
    assert openpyxl.load_workbook(out).sheetnames == ["Data"]


# ---------------------------------------------------------------------------
# naming many sheets at once
# ---------------------------------------------------------------------------

@pytest.fixture
def excel_with_many_tabs(tmp_path: Path) -> Path:
    """Three numbered tabs of one shape, a data sheet, and an empty divider."""
    wb = openpyxl.Workbook()
    wb.active.title = "Data"
    wb.active.cell(1, 1, "name")
    wb.active.cell(2, 1, "Smith")
    for number in ("48", "49", "77"):
        tab = wb.create_sheet(number)
        tab.cell(1, 1, "comment")
        tab.cell(2, 1, f"note {number}")
    wb.create_sheet("AC-DC")
    path = tmp_path / "book.xlsx"
    wb.save(path)
    return path


def test_a_pattern_reaches_every_matching_sheet(excel_with_many_tabs, tmp_path):
    """TF1 has 55 cashflow tabs in production and TF8 over a hundred."""
    config = write_config(
        tmp_path,
        {
            "keep_sheets": ["Data", "AC-DC"],
            "groups": [
                {
                    "name": "comments",
                    "prefix": "TEXT",
                    "columns": [{"sheet_pattern": r"\d+", "col": "A", "data_from_row": 2}],
                }
            ],
        },
    )
    anonymize(excel_with_many_tabs, config)

    out = openpyxl.load_workbook(excel_with_many_tabs.with_stem(excel_with_many_tabs.stem + "_anonymized"))
    assert [out[tab].cell(2, 1).value.startswith("TEXT") for tab in ("48", "49", "77")] == [True] * 3
    assert out["Data"].cell(2, 1).value == "Smith", "the pattern must not reach beyond it"


def test_a_pattern_accounts_for_the_sheets_it_matches(excel_with_many_tabs, tmp_path):
    """Otherwise every tab would have to be listed twice."""
    config = write_config(
        tmp_path,
        {
            "unlisted_sheets": "fail",
            "keep_sheets": ["Data", "AC-DC"],
            "groups": [
                {
                    "name": "comments",
                    "prefix": "TEXT",
                    "columns": [{"sheet_pattern": r"\d+", "col": "A", "data_from_row": 2}],
                }
            ],
        },
    )
    anonymize(excel_with_many_tabs, config)


def test_empty_sheets_can_be_removed_as_a_class(excel_with_many_tabs, tmp_path):
    """The dividers are empty and some of them are named after a transaction."""
    config = write_config(
        tmp_path,
        {
            "ignore_sheets": [{"empty": True, "reason": "section dividers"}],
            "keep_sheets": ["Data"],
            "groups": [
                {
                    "name": "comments",
                    "prefix": "TEXT",
                    "columns": [{"sheet_pattern": r"\d+", "col": "A", "data_from_row": 2}],
                }
            ],
        },
    )
    anonymize(excel_with_many_tabs, config)

    out = openpyxl.load_workbook(excel_with_many_tabs.with_stem(excel_with_many_tabs.stem + "_anonymized"))
    assert "AC-DC" not in out.sheetnames
    assert set(out.sheetnames) == {"Data", "48", "49", "77"}


def test_a_pattern_that_matches_nothing_stops_the_run(excel_with_many_tabs, tmp_path):
    """The one mistake no other check catches.

    `'\\d+'` in single-quoted YAML is a literal backslash followed by a
    `d`, so it matches no sheet at all. The columns it named went
    unreplaced, and the configuration, the structural comparison and the
    residual check all reported the file as sound.
    """
    config = write_config(
        tmp_path,
        {
            "keep_sheets": ["Data", "AC-DC"],
            "groups": [
                {
                    "name": "comments",
                    "prefix": "TEXT",
                    "columns": [{"sheet_pattern": r"\\d+", "col": "A", "data_from_row": 2}],
                }
            ],
        },
    )
    with pytest.raises(ValueError, match="matches no sheet"):
        anonymize(excel_with_many_tabs, config)


def test_a_named_sheet_that_is_absent_from_a_group_stops_the_run(excel_with_many_tabs, tmp_path):
    """A sheet renamed between exports leaves its column unhandled."""
    config = write_config(
        tmp_path,
        {
            "keep_sheets": ["Data", "AC-DC"],
            "groups": [
                {
                    "name": "names",
                    "prefix": "NAME",
                    "columns": [{"sheet": "Daten", "col": "A", "data_from_row": 2}],
                }
            ],
        },
    )
    with pytest.raises(ValueError, match="matches no sheet"):
        anonymize(excel_with_many_tabs, config)


# ---------------------------------------------------------------------------
# addressing a range of columns
# ---------------------------------------------------------------------------

@pytest.fixture
def excel_with_a_scratch_area(tmp_path: Path) -> Path:
    """Data in A and B, and helper columns to the right of it.

    The helper columns sit at different letters in each tab, which is
    why they cannot be listed by letter — the cashflow tabs of the
    DataSet workbooks are exactly this shape.
    """
    wb = openpyxl.Workbook()
    first = wb.active
    first.title = "1"
    first.append(["id", "amount", "sum", "note"])
    first.append([1, 1000, 1000, "Miller"])
    second = wb.create_sheet("2")
    second.append(["id", "amount", "spacer", "sum", "note"])
    second.append([2, 2000, None, 2000, "Miller"])
    path = tmp_path / "book.xlsx"
    wb.save(path)
    return path


def test_a_range_reaches_the_helper_columns_wherever_they_sit(excel_with_a_scratch_area, tmp_path):
    """The originals hide in the columns beyond the data."""
    config = write_config(
        tmp_path,
        {
            "groups": [
                {
                    "name": "amounts",
                    "strategy": "scale",
                    "factor": 2,
                    "columns": [{"sheet_pattern": r"\d+", "col_from": "B", "data_from_row": 2}],
                },
                {
                    "name": "text",
                    "prefix": "TEXT",
                    "columns": [{"sheet_pattern": r"\d+", "col_from": "B", "data_from_row": 2}],
                },
            ]
        },
    )
    anonymize(excel_with_a_scratch_area, config)

    out = openpyxl.load_workbook(excel_with_a_scratch_area.with_stem(excel_with_a_scratch_area.stem + "_anonymized"))
    assert [out["1"].cell(2, c).value for c in (2, 3)] == [2000, 2000]
    assert [out["2"].cell(2, c).value for c in (2, 4)] == [4000, 4000]
    assert out["1"].cell(2, 4).value.startswith("TEXT"), "the note is in D here"
    assert out["2"].cell(2, 5).value.startswith("TEXT"), "and in E there"
    assert out["1"].cell(2, 1).value == 1, "nothing left of the range is touched"


def test_a_range_can_be_bounded_on_the_right(excel_with_a_scratch_area, tmp_path):
    config = write_config(
        tmp_path,
        {
            "groups": [
                {
                    "name": "amounts",
                    "strategy": "scale",
                    "factor": 2,
                    "columns": [
                        {"sheet": "1", "col_from": "B", "col_to": "C", "data_from_row": 2},
                    ],
                }
            ]
        },
    )
    anonymize(excel_with_a_scratch_area, config)

    out = openpyxl.load_workbook(excel_with_a_scratch_area.with_stem(excel_with_a_scratch_area.stem + "_anonymized"))
    assert [out["1"].cell(2, c).value for c in (2, 3)] == [2000, 2000]
    assert out["1"].cell(2, 4).value == "Miller", "D is beyond col_to"


def test_a_key_group_leaves_dates_alone(tmp_path):
    """A key over a date turns a date column into a text column.

    A range of columns meets dates whether it meant to or not — the
    scratch area of a cashflow tab repeats the cashflow date four times.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["name", "when"])
    ws.append(["Miller", datetime.date(2024, 3, 1)])
    path = tmp_path / "book.xlsx"
    wb.save(path)

    config = write_config(
        tmp_path,
        {
            "groups": [
                {
                    "name": "everything",
                    "prefix": "X",
                    "columns": [{"sheet": "Data", "col_from": "A", "data_from_row": 2}],
                }
            ]
        },
    )
    anonymize(path, config)

    out = openpyxl.load_workbook(path.with_stem(path.stem + "_anonymized"))["Data"]
    assert out.cell(2, 1).value.startswith("X"), "the name is replaced"
    assert out.cell(2, 2).value == datetime.datetime(2024, 3, 1), "the date is not"


def test_a_query_table_becomes_a_plain_table(tmp_path):
    """Excel offers to repair a table whose query is not in the file.

    openpyxl carries neither the query nor the connection over, so the
    declaration has to go with them.
    """
    from openpyxl.worksheet.table import Table, TableColumn

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["rate", "ccy"])
    ws.append([1.0, "EUR"])
    table = Table(
        displayName="Series3",
        ref="A1:B2",
        tableType="queryTable",
        tableColumns=[
            TableColumn(id=1, name="rate", queryTableFieldId=1),
            TableColumn(id=2, name="ccy", queryTableFieldId=2),
        ],
    )
    ws.add_table(table)
    path = tmp_path / "book.xlsx"
    wb.save(path)

    config = write_config(tmp_path, {"keep_sheets": ["Data"], "groups": []})
    anonymize(path, config)

    out = openpyxl.load_workbook(path.with_stem(path.stem + "_anonymized"))["Data"]
    written = out.tables["Series3"]
    assert written.tableType is None
    assert [column.queryTableFieldId for column in written.tableColumns] == [None, None]


def test_replacing_a_value_removes_the_link_under_it(tmp_path):
    """A hyperlink is a relationship of the sheet, not a value in the cell.

    A column of email addresses replaced by keys kept the addresses,
    where no comparison of cell values would ever have looked.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["contact"])
    ws.cell(2, 1, "a.miller@example.com").hyperlink = "mailto:a.miller@example.com"
    path = tmp_path / "book.xlsx"
    wb.save(path)

    config = write_config(
        tmp_path,
        {
            "groups": [
                {"name": "people", "prefix": "PERSON", "columns": [{"sheet": "Data", "col": "A", "data_from_row": 2}]}
            ]
        },
    )
    anonymize(path, config)

    out_path = path.with_stem(path.stem + "_anonymized")
    out = openpyxl.load_workbook(out_path)["Data"]
    assert out.cell(2, 1).value.startswith("PERSON")
    assert out.cell(2, 1).hyperlink is None
    with zipfile.ZipFile(out_path) as archive:
        stored = b"".join(archive.read(name) for name in archive.namelist())
    assert b"a.miller@example.com" not in stored, "not anywhere in the file, not only in the cell"


def test_cell_comments_are_removed(tmp_path):
    """Comments are prose, and prose cannot be checked for names."""
    from openpyxl.comments import Comment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["amount"])
    ws.cell(2, 1, 1000).comment = Comment("Follow up with Miller & Co.", "Anna")
    path = tmp_path / "book.xlsx"
    wb.save(path)

    config = write_config(tmp_path, {"keep_sheets": ["Data"], "groups": []})
    anonymize(path, config)

    out_path = path.with_stem(path.stem + "_anonymized")
    assert openpyxl.load_workbook(out_path)["Data"].cell(2, 1).comment is None
    with zipfile.ZipFile(out_path) as archive:
        stored = b"".join(archive.read(name) for name in archive.namelist())
    assert b"Miller" not in stored and b"Anna" not in stored


def test_comments_can_be_kept_on_purpose(tmp_path):
    from openpyxl.comments import Comment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["amount"])
    ws.cell(2, 1, 1000).comment = Comment("a note", "Anna")
    path = tmp_path / "book.xlsx"
    wb.save(path)

    config = write_config(tmp_path, {"comments": "keep", "keep_sheets": ["Data"], "groups": []})
    anonymize(path, config)

    out = openpyxl.load_workbook(path.with_stem(path.stem + "_anonymized"))["Data"]
    assert out.cell(2, 1).comment is not None


def test_the_document_properties_are_cleared(tmp_path):
    """A workbook names the people who saved it, outside every cell."""
    from openpyxl.packaging.custom import StringProperty

    wb = openpyxl.Workbook()
    wb.active.title = "Data"
    wb.active.append(["amount"])
    wb.properties.creator = "Angela Example"
    wb.properties.lastModifiedBy = "Thorsten Example"
    wb.properties.title = "TF6.xlsx"
    wb.custom_doc_props.append(StringProperty(name="Label_Owner", value="p.benedikt@example.com"))
    path = tmp_path / "book.xlsx"
    wb.save(path)

    config = write_config(tmp_path, {"keep_sheets": ["Data"], "groups": []})
    anonymize(path, config)

    out_path = path.with_stem(path.stem + "_anonymized")
    with zipfile.ZipFile(out_path) as archive:
        stored = b"".join(archive.read(name) for name in archive.namelist())
    for gone in (b"Angela Example", b"Thorsten Example", b"p.benedikt@example.com", b"TF6.xlsx"):
        assert gone not in stored, gone


# ---------------------------------------------------------------------------
# a pattern is a bet about the layout
# ---------------------------------------------------------------------------

@pytest.fixture
def excel_with_a_tab_out_of_line(tmp_path: Path) -> Path:
    """Three numbered tabs, one of which has a column inserted."""
    wb = openpyxl.Workbook()
    wb.active.title = "1"
    wb.active.append(["investment_id", "amount", "comment"])
    wb.active.append([1, 1000, "note"])
    second = wb.create_sheet("2")
    second.append(["investment_id", "amount", "comment"])
    second.append([2, 2000, "note"])
    third = wb.create_sheet("3")
    third.append(["investment_id", "PIK comment", "amount"])
    third.append([3, "note", 3000])
    path = tmp_path / "book.xlsx"
    wb.save(path)
    return path


def test_a_pattern_that_reaches_two_layouts_is_reported(excel_with_a_tab_out_of_line, tmp_path, capsys):
    """Otherwise the letters land on the neighbouring column in silence."""
    config = write_config(
        tmp_path,
        {
            "groups": [
                {
                    "name": "amounts",
                    "strategy": "scale",
                    "factor": 2,
                    "columns": [{"sheet_pattern": r"\d+", "col": "B", "data_from_row": 2}],
                }
            ]
        },
    )
    anonymize(excel_with_a_tab_out_of_line, config)

    printed = capsys.readouterr().out
    assert "1 of 3 sheets matched by a pattern do not have the header the letters assume" in printed
    assert "in column B" in printed
    listing = excel_with_a_tab_out_of_line.with_name(
        f"{excel_with_a_tab_out_of_line.stem}_differing_sheets.txt"
    )
    assert listing.read_text(encoding="utf-8").split() == ["3"]


def test_fail_refuses_a_pattern_that_reaches_two_layouts(excel_with_a_tab_out_of_line, tmp_path):
    config = write_config(
        tmp_path,
        {
            "unlisted_sheets": "fail",
            "groups": [
                {
                    "name": "amounts",
                    "strategy": "scale",
                    "factor": 2,
                    "columns": [{"sheet_pattern": r"\d+", "col": "B", "data_from_row": 2}],
                }
            ],
        },
    )
    with pytest.raises(ValueError, match="do not have the layout their pattern assumes"):
        anonymize(excel_with_a_tab_out_of_line, config)


def test_the_scratch_area_beyond_the_named_columns_may_differ(excel_with_a_tab_out_of_line, tmp_path, capsys):
    """A range is addressed precisely because those columns vary."""
    config = write_config(
        tmp_path,
        {
            "groups": [
                {
                    "name": "ids",
                    "prefix": "ID",
                    "columns": [{"sheet_pattern": r"\d+", "col": "A", "data_from_row": 2}],
                }
            ]
        },
    )
    anonymize(excel_with_a_tab_out_of_line, config)

    assert "differ in their header row" not in capsys.readouterr().out


# ---------------------------------------------------------------------------
# What a workbook keeps about other workbooks
# ---------------------------------------------------------------------------

def _with_a_link_to_another_workbook(path: Path) -> None:
    """Give the workbook at `path` one external link, cache and all.

    A link records three things and every one of them is a leak: where
    the other file was, what its ranges are called, and the values last
    read from it. The PE tracker carries thirty such links.
    """
    from openpyxl.packaging.relationship import Relationship
    from openpyxl.workbook.external_link.external import (
        ExternalBook, ExternalCell, ExternalDefinedName, ExternalLink,
        ExternalRow, ExternalSheetData, ExternalSheetDataSet, ExternalSheetNames,
    )

    wb = openpyxl.load_workbook(path)
    book = ExternalBook(
        sheetNames=ExternalSheetNames(sheetName=["Rates"]),
        definedNames=[ExternalDefinedName(name="Gamma_Commitment", refersTo="Rates!$A$1")],
        sheetDataSet=ExternalSheetDataSet(sheetData=[
            ExternalSheetData(sheetId=0, row=[
                ExternalRow(r=1, cell=[ExternalCell(r="A1", t="str", v="Gamma Holdings AG")])
            ])
        ]),
    )
    link = ExternalLink(externalBook=book)
    link.file_link = Relationship(
        Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLinkPath",
        Target="/Users/someone/Desktop/Rates.xlsx",
        TargetMode="External",
    )
    wb._external_links = [link]
    wb.save(path)


def test_links_to_other_workbooks_are_removed(tmp_path):
    """Nothing in the run would ever have looked at them: the cells being
    replaced are in this workbook, and a cache is not a cell."""
    wb = openpyxl.Workbook()
    wb.active.title = "Data"
    wb.active.append(["contact"])
    wb.active.append(["Smith"])
    path = tmp_path / "book.xlsx"
    wb.save(path)
    _with_a_link_to_another_workbook(path)

    config = write_config(
        tmp_path,
        {"groups": [{"name": "people", "prefix": "PERSON",
                     "columns": [{"sheet": "Data", "col": "A", "data_from_row": 2}]}]},
    )
    anonymize(path, config)

    out_path = path.with_stem(path.stem + "_anonymized")
    assert not openpyxl.load_workbook(out_path)._external_links
    with zipfile.ZipFile(out_path) as archive:
        stored = b"".join(archive.read(name) for name in archive.namelist())
    assert b"Gamma Holdings AG" not in stored, "the value it had cached"
    assert b"Gamma_Commitment" not in stored, "what the other workbook called the range"
    assert b"someone/Desktop" not in stored, "where that workbook stood, and whose desktop it was"


def test_named_ranges_go_with_the_formulas(tmp_path):
    """A name is written by hand and says what the range is about. Once
    the formulas are resolved nothing reads it any more."""
    wb = openpyxl.Workbook()
    wb.active.title = "Data"
    wb.active.append(["amount"])
    wb.active.append([1000])
    wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName("Gamma_Commitment", attr_text="Data!$A$2"))
    path = tmp_path / "book.xlsx"
    wb.save(path)

    config = write_config(tmp_path, {"keep_sheets": ["Data"], "groups": []})
    anonymize(path, config)

    out_path = path.with_stem(path.stem + "_anonymized")
    assert not openpyxl.load_workbook(out_path).defined_names
    with zipfile.ZipFile(out_path) as archive:
        stored = b"".join(archive.read(name) for name in archive.namelist())
    assert b"Gamma_Commitment" not in stored


def test_named_ranges_stay_where_the_formulas_do(tmp_path):
    """In `keep` mode removing them would break every formula that uses one."""
    wb = openpyxl.Workbook()
    wb.active.title = "Data"
    wb.active.append(["amount"])
    wb.active.append([1000])
    wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName("Commitment", attr_text="Data!$A$2"))
    path = tmp_path / "book.xlsx"
    wb.save(path)

    config = write_config(tmp_path, {"formulas": "keep", "keep_sheets": ["Data"], "groups": []})
    anonymize(path, config)

    out_path = path.with_stem(path.stem + "_anonymized")
    assert "Commitment" in openpyxl.load_workbook(out_path).defined_names


def test_pivot_tables_go_with_their_cached_source(tmp_path):
    """A pivot table draws from a copy of the source range stored beside
    it, not from the sheet. Anonymising the sheet leaves that copy."""
    from anonymize import drop_pivot_caches

    wb = openpyxl.Workbook()
    wb._pivots = ["cache of the original rows"]
    wb.active._pivots = ["the table drawn from it"]

    drop_pivot_caches(wb)

    assert wb._pivots == []
    assert wb.active._pivots == []
