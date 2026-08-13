#!/usr/bin/env python3

import argparse
import json
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


def cell_info(cell, include_values=False):
    info = {
        "address": cell.coordinate,
        "data_type": cell.data_type,
    }

    if cell.number_format:
        info["number_format"] = cell.number_format

    if cell.style_id:
        info["style_id"] = cell.style_id

    if cell.comment:
        info["comment"] = True

    if cell.hyperlink:
        info["hyperlink"] = True

    if include_values:
        info["value"] = cell.value

    return info


def worksheet_info(ws, results=None, header_row=None,
                   include_values=False, include_cells=False):
    info = {
        "name": ws.title,
        "state": ws.sheet_state,
        "dimensions": {
            "min_row": ws.min_row,
            "max_row": ws.max_row,
            "min_column": ws.min_column,
            "max_column": ws.max_column,
            "range": ws.calculate_dimension(),
        },
        "freeze_panes": (
            ws.freeze_panes.coordinate
            if ws.freeze_panes is not None
            and hasattr(ws.freeze_panes, "coordinate")
            else ws.freeze_panes
        ),
    }

    if header_row:
        info["header_row"] = header_row

    info["columns"] = column_analysis(
        ws,
        results=results,
        header_row=header_row,
        include_examples=include_values,
    )

    # Hidden rows
    hidden_rows = [
        row
        for row, dimension in ws.row_dimensions.items()
        if dimension.hidden
    ]

    if hidden_rows:
        info["hidden_rows"] = hidden_rows

    # Hidden columns
    hidden_columns = []

    for column, dimension in ws.column_dimensions.items():
        if dimension.hidden:
            hidden_columns.append(column)

    if hidden_columns:
        info["hidden_columns"] = hidden_columns

    # Merged cells
    merged = [str(rng) for rng in ws.merged_cells.ranges]

    if merged:
        info["merged_cells"] = merged

    # AutoFilter
    if ws.auto_filter and ws.auto_filter.ref:
        info["auto_filter"] = ws.auto_filter.ref

    # Excel Tables
    tables = []

    for table_name in ws.tables:
        table = ws.tables[table_name]

        tables.append({
            "name": table_name,
            "range": table.ref,
        })

    if tables:
        info["tables"] = tables
        tables = []

    
    # Data validation
    validations = []

    if ws.data_validations:
        for dv in ws.data_validations.dataValidation:
            validations.append({
                "type": dv.type,
                "operator": dv.operator,
                "formula1": dv.formula1,
                "formula2": dv.formula2,
                "ranges": str(dv.sqref),
            })

    if validations:
        info["data_validations"] = validations

    # Conditional formatting
    conditional_formatting = []

    for cf in ws.conditional_formatting:
        conditional_formatting.append(str(cf))

    if conditional_formatting:
        info["conditional_formatting"] = conditional_formatting

    # Cells, one entry each. Off by default: the FX sheet of one
    # workbook holds 108,532 rows, and a manifest with a line per cell
    # cannot be read by anyone.
    if include_cells:
        cells = []

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None or cell.has_style or cell.comment:
                    cells.append(
                        cell_info(
                            cell,
                            include_values=include_values,
                        )
                    )

        info["cell_count"] = len(cells)

        if cells:
            info["cells"] = cells

    return info

def value_type(cell, value):
    if cell.is_date:
        return "date"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "number"
    if isinstance(value, str):
        return "string"

    return type(value).__name__


def column_analysis(ws, results=None, header_row=None,
                    include_examples=False, example_count=3):
    # Two questions, two answers, and they are not the same one.
    #
    # For the migration a formula cell is derived and is not taken over:
    # the target model recomputes it. `formula_count` beside
    # `value_count` is what makes that a decision rather than an
    # assumption — a column where the two differ is part formula and part
    # typed in, and the typed-in cells are original data.
    #
    # For the anonymisation it does not matter that a formula stood
    # there. `formulas: values` resolves every one of them, and what
    # remains is the real figure. So the types reported here are the
    # types of the *results*, read from a second copy of the workbook
    # opened with `data_only=True`. Reading the formulas alone reports a
    # workbook of derived data as one formula column after another.
    columns = {}
    first_row = header_row + 1 if header_row else ws.min_row

    for col_idx in range(ws.min_column, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)

        values = []
        type_counts = {}
        formula_count = 0

        for row_idx in range(first_row, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)
            result = results.cell(row_idx, col_idx) if results else cell

            if cell.data_type == "f":
                formula_count += 1

            value = result.value

            if value is None:
                continue

            values.append(value)

            kind = value_type(result, value)

            type_counts[kind] = (
                type_counts.get(kind, 0) + 1
            )

        total_rows = ws.max_row - first_row + 1
        value_count = len(values)

        info = {
            "column": column_letter,
            "value_count": value_count,
            "empty_count": total_rows - value_count,
            "unique_count": len(set(values)),
            "types": type_counts,
        }

        if header_row:
            heading = ws.cell(header_row, col_idx).value

            info["header"] = (
                str(heading).strip()
                if heading is not None
                else ""
            )

        if formula_count:
            info["formula_count"] = formula_count

        if value_count:
            if len(type_counts) == 1:
                info["type"] = next(iter(type_counts))
            else:
                info["type"] = "mixed"

        if include_examples:
            info["examples"] = [
                str(value)
                for value in values[:example_count]
            ]

        if values and all(
            isinstance(v, (int, float))
            and not isinstance(v, bool)
            for v in values
        ):
            info["min"] = min(values)
            info["max"] = max(values)

        columns[column_letter] = info

    return columns


def workbook_info(filename, include_values=False, include_cells=False,
                  header_rows=None):
    wb = openpyxl.load_workbook(
        filename,
        data_only=False,
        read_only=False,
    )

    # The same workbook a second time, with the results Excel last
    # calculated in place of the formulas. openpyxl gives one or the
    # other and never both.
    values = openpyxl.load_workbook(
        filename,
        data_only=True,
        read_only=False,
    )

    header_rows = header_rows or {}

    manifest = {
        "format": Path(filename).suffix.lstrip(".").lower(),
        "file": Path(filename).name,
        "worksheets": [],
    }

    # Workbook properties
    if wb.properties:
        props = {}

        for attr in (
            "created",
            "modified",
            "title",
            "subject",
            "description",
            "category",
        ):
            value = getattr(wb.properties, attr, None)

            if value is not None:
                props[attr] = str(value)

        if props:
            manifest["properties"] = props

    # Active sheet
    try:
        manifest["active_worksheet"] = wb.active.title
    except Exception:
        pass

    # Defined names
    defined_names = []

    for name, defined_name in wb.defined_names.items():
        defined_names.append({
            "name": name,
            "value": str(defined_name.attr_text),
            "local_sheet_id": defined_name.localSheetId,
        })

    if defined_names:
        manifest["defined_names"] = defined_names

    # Worksheets
    for ws in wb.worksheets:
        header_row = header_rows.get(ws.title, header_rows.get("*"))

        manifest["worksheets"].append(
            worksheet_info(
                ws,
                results=values[ws.title],
                header_row=header_row,
                include_values=include_values,
                include_cells=include_cells,
            )
        )

    return manifest


def parse_header_rows(arguments):
    # Either one row for every sheet (`--header-row 3`) or one per sheet
    # (`--header-row "Cash Transactions=3"`). The workbooks this was
    # written for keep a title, a block description and a column
    # description above the headings, and not the same number of them on
    # every sheet.
    header_rows = {}

    for argument in arguments or []:
        sheet, _, row = argument.rpartition("=")

        if not row.strip().isdigit():
            raise SystemExit(
                f"ERROR: --header-row expects N or SHEET=N, not {argument!r}"
            )

        header_rows[sheet or "*"] = int(row)

    return header_rows


def main():
    parser = argparse.ArgumentParser(
        description="Inspect the structure of an Excel file."
    )

    parser.add_argument(
        "file",
        type=Path,
        help="Excel XLSX file",
    )

    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Output JSON file",
    )

    parser.add_argument(
        "--values",
        action="store_true",
        help="Include cell values in the manifest",
    )

    parser.add_argument(
        "--cells",
        action="store_true",
        help="Include one entry per cell (large)",
    )

    parser.add_argument(
        "--header-row",
        action="append",
        metavar="N | SHEET=N",
        help="Row holding the column headings, for all sheets or for one",
    )

    args = parser.parse_args()

    if not args.file.exists():
        raise SystemExit(
            f"ERROR: File not found: {args.file}"
        )

    # .xlsm is .xlsx with macros beside it. openpyxl reads it the same
    # way and drops the macros on save, which is what a derived dataset
    # wants anyway — the data is being taken over, not the workbook.
    if args.file.suffix.lower() not in (".xlsx", ".xlsm"):
        raise SystemExit(
            "ERROR: Only XLSX and XLSM files are supported."
        )

    print(f"Inspecting: {args.file}")

    manifest = workbook_info(
        args.file,
        include_values=args.values,
        include_cells=args.cells,
        header_rows=parse_header_rows(args.header_row),
    )

    output = json.dumps(
        manifest,
        indent=2,
        ensure_ascii=False,
        default=str,
    )

    if args.output:
        args.output.write_text(
            output,
            encoding="utf-8",
        )

        print(f"Manifest written to: {args.output}")

    else:
        print(output)


if __name__ == "__main__":
    main()
