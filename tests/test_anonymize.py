"""Tests for anonymize.py: key-based replacement."""

import json
import openpyxl
import pytest
from pathlib import Path

from anonymize import anonymize


def _load_output(excel_path: Path, suffix: str) -> openpyxl.worksheet.worksheet.Worksheet:
    out = excel_path.with_stem(excel_path.stem + suffix)
    return openpyxl.load_workbook(out)["Sheet1"]


# ---------------------------------------------------------------------------
# Basic replacement
# ---------------------------------------------------------------------------

def test_original_values_are_gone(sample_excel, anon_config):
    anonymize(sample_excel, anon_config)
    ws = _load_output(sample_excel, "_anonymized")
    data_values = [ws.cell(r, c).value for r in (2, 3) for c in (1, 2, 3, 4)]
    assert "Smith" not in data_values
    assert "Brown" not in data_values
    assert "John"  not in data_values
    assert "Jane"  not in data_values


def test_keys_have_correct_prefix(sample_excel, anon_config):
    anonymize(sample_excel, anon_config)
    ws = _load_output(sample_excel, "_anonymized")
    assert ws.cell(2, 1).value.startswith("LAST")
    assert ws.cell(2, 2).value.startswith("FIRST")


# ---------------------------------------------------------------------------
# Relationship preservation
# ---------------------------------------------------------------------------

def test_same_value_gets_same_key_across_columns(sample_excel, anon_config):
    """Smith in col A (employee) and col C (manager) must get the same key."""
    anonymize(sample_excel, anon_config)
    ws = _load_output(sample_excel, "_anonymized")
    # A2 = "Smith" key, C3 = "Smith" key → must be equal
    assert ws.cell(2, 1).value == ws.cell(3, 3).value
    # A3 = "Brown" key, C2 = "Brown" key → must be equal
    assert ws.cell(3, 1).value == ws.cell(2, 3).value


def test_different_values_get_different_keys(sample_excel, anon_config):
    anonymize(sample_excel, anon_config)
    ws = _load_output(sample_excel, "_anonymized")
    assert ws.cell(2, 1).value != ws.cell(3, 1).value  # Smith ≠ Brown


def test_groups_are_independent(sample_excel, anon_config):
    """last_names and first_names groups have separate key spaces."""
    anonymize(sample_excel, anon_config)
    ws = _load_output(sample_excel, "_anonymized")
    last_keys  = {ws.cell(r, c).value for r in (2, 3) for c in (1, 3)}
    first_keys = {ws.cell(r, c).value for r in (2, 3) for c in (2, 4)}
    assert last_keys.isdisjoint(first_keys), "Key spaces of different groups must not overlap"


# ---------------------------------------------------------------------------
# Header rows untouched
# ---------------------------------------------------------------------------

def test_header_row_is_not_anonymized(sample_excel, anon_config):
    anonymize(sample_excel, anon_config)
    ws = _load_output(sample_excel, "_anonymized")
    assert ws.cell(1, 1).value == "Last Name"
    assert ws.cell(1, 2).value == "First Name"
    assert ws.cell(1, 3).value == "Mgr Last"
    assert ws.cell(1, 4).value == "Mgr First"


# ---------------------------------------------------------------------------
# Mapping file
# ---------------------------------------------------------------------------

def test_mapping_file_is_saved(sample_excel, anon_config):
    anonymize(sample_excel, anon_config)
    map_path = sample_excel.parent / "anon_map.json"
    assert map_path.exists()


def test_mapping_contains_original_values(sample_excel, anon_config):
    anonymize(sample_excel, anon_config)
    mapping = json.loads((sample_excel.parent / "anon_map.json").read_text())
    assert "Smith" in mapping["last_names"]
    assert "Brown" in mapping["last_names"]
    assert "John"  in mapping["first_names"]
    assert "Jane"  in mapping["first_names"]


def test_mapping_values_match_output(sample_excel, anon_config):
    """Keys in the mapping must equal what's written in the output file."""
    anonymize(sample_excel, anon_config)
    mapping = json.loads((sample_excel.parent / "anon_map.json").read_text())
    ws = _load_output(sample_excel, "_anonymized")
    assert ws.cell(2, 1).value == mapping["last_names"]["Smith"]
    assert ws.cell(3, 1).value == mapping["last_names"]["Brown"]


# ---------------------------------------------------------------------------
# Roundtrip: anonymize → deanonymize
# ---------------------------------------------------------------------------

def test_roundtrip_restores_original_values(sample_excel, anon_config):
    from deanonymize import deanonymize

    anonymize(sample_excel, anon_config)
    anon_path = sample_excel.with_stem(sample_excel.stem + "_anonymized")
    map_path  = sample_excel.parent / "anon_map.json"

    deanonymize(anon_path, map_path)
    restored_path = anon_path.with_stem(anon_path.stem + "_restored")

    original = openpyxl.load_workbook(sample_excel)["Sheet1"]
    restored = openpyxl.load_workbook(restored_path)["Sheet1"]

    for row in range(1, 4):
        for col in range(1, 5):
            assert original.cell(row, col).value == restored.cell(row, col).value
