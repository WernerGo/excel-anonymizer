"""
Excel anonymization via key replacement.

Replaces configured cell values with unique keys. Identical values within
a group always get the same key, preserving relationships between columns
(e.g. employee → manager links stay intact in the anonymized file).

Numeric columns can be scaled by a factor instead of replaced, so that
amounts become fictitious while the relationships between them survive.

Usage:
    python src/anonymize.py input.xlsx
    python src/anonymize.py input.xlsx --config my_config.yaml
    python src/anonymize.py input.xlsx --config examples/config_names.yaml
"""

import argparse
import datetime
import json
import re
import sys
import zipfile
from collections import OrderedDict
from pathlib import Path

import openpyxl
import yaml

FORMULA_MODES = ("values", "keep")


def load_source(excel_path: Path, mode: str) -> openpyxl.Workbook:
    """Open the workbook, resolving formulas or keeping them.

    An .xlsx stores two things for a formula cell: the formula, and the
    result Excel last calculated for it. openpyxl loads one or the other,
    and writes back whatever it loaded. So the choice is real and it is
    lossy either way:

    ``values``
        Formulas are replaced by their last results. The output is a
        plain value workbook — what a program that reads values wants,
        and reproducible because nothing recalculates.
    ``keep``
        Formulas survive, their results do not. The file has to be
        opened and recalculated in Excel before anything can read a
        value from it.

    Args:
        excel_path: Path to the input Excel file (.xlsx).
        mode: ``values`` or ``keep``.

    Returns:
        The loaded workbook.
    """
    if mode not in FORMULA_MODES:
        raise ValueError(f"formulas: {mode!r} — expected one of {', '.join(FORMULA_MODES)}")
    if mode == "keep":
        print("  WARNING: formulas are kept, so their last results are dropped.")
        print("           Open the output in Excel and recalculate before reading values from it.")
    return openpyxl.load_workbook(excel_path, data_only=(mode == "values"))


CELL = re.compile(rb"<c\b[^>]*>(?:(?!</c>).)*</c>", re.S)


def count_uncached_formulas(excel_path: Path) -> int:
    """Count formula cells that were never calculated.

    Not the same as a formula whose result is empty. ``=IFERROR(…, "")``
    is calculated and its answer is nothing, which the file records as an
    empty value element — the cell is correct and the output should be
    empty there too. A cell that was never calculated has no value
    element at all, and in ``values`` mode it arrives empty without
    anything saying it ever held something.

    openpyxl reports both as None, so this reads the stored XML, where
    the two are plainly different.

    Args:
        excel_path: Path to the input Excel file (.xlsx).

    Returns:
        The number of formula cells that carry no result of any kind.
    """
    missing = 0
    with zipfile.ZipFile(excel_path) as archive:
        sheets = [name for name in archive.namelist() if name.startswith("xl/worksheets/sheet")]
        for sheet in sheets:
            for match in CELL.finditer(archive.read(sheet)):
                cell = match.group(0)
                if b"<f" in cell and b"<v" not in cell and b"<is" not in cell:
                    missing += 1
    return missing


def sheet_is_empty(ws) -> bool:
    """Whether a worksheet holds no value in any cell.

    The DataSet workbooks use empty sheets as section dividers, and some
    of them are named after a transaction. They carry no data, but a
    name is a value too.
    """
    return not any(cell.value is not None for row in ws.iter_rows() for cell in row)


def matching_sheets(wb: openpyxl.Workbook, spec, warn_if_absent: bool = True) -> list[str]:
    """Return the sheets one entry refers to.

    An entry names a sheet exactly (``sheet``), by a regular expression
    over the whole name (``sheet_pattern``), or by emptiness
    (``empty: true``). The pattern is what makes the cashflow tabs
    manageable: TF1 has 55 of them in production and TF8 over a hundred,
    all with the same structure and a number for a name.

    Args:
        wb: The open workbook.
        spec: A string, or a mapping with one of the three keys.
        warn_if_absent: Whether to report a named sheet that is not there.

    Returns:
        The matching sheet names, in workbook order.
    """
    if not isinstance(spec, dict):
        spec = {"sheet": spec}

    if spec.get("empty"):
        return [name for name in wb.sheetnames if sheet_is_empty(wb[name])]

    if "sheet_pattern" in spec:
        pattern = re.compile(spec["sheet_pattern"])
        return [name for name in wb.sheetnames if pattern.fullmatch(name)]

    name = str(spec["sheet"])
    if name in wb.sheetnames:
        return [name]
    if warn_if_absent:
        print(f"  WARNING: sheet '{name}' is named in the config but not in the file.")
    return []


def drop_sheets(wb: openpyxl.Workbook, entries: list) -> None:
    """Remove sheets that are not part of what is being passed on.

    A workbook usually carries more than the data someone needs from it:
    working views, report layouts, exports to another system. Those
    sheets hold the same values as the data sheets, so anonymizing the
    data sheets alone leaves the file full of originals. Removing them
    is both simpler and more complete than replacing them column by
    column.

    Each entry names a sheet and, where it is written that way, the
    reason it can go — the list is meant to be read by whoever has to
    agree that these sheets are dispensable.

    Args:
        wb: The open workbook.
        entries: Sheet names, or mappings with ``sheet`` and ``reason``.
    """
    names: list[str] = []
    for entry in entries:
        names.extend(name for name in matching_sheets(wb, entry) if name not in names)

    if names and not set(wb.sheetnames) - set(names):
        raise ValueError("ignore_sheets would remove every sheet")

    for name in names:
        del wb[name]

    removed = list(names)
    if removed:
        print(f"  Removed {len(removed)} sheets that are not part of the delivery.")
        drop_dangling_names(wb, removed)


def drop_dangling_names(wb: openpyxl.Workbook, removed: list[str]) -> None:
    """Remove defined names that point at a sheet that is no longer there.

    Deleting a sheet leaves the names that referred to it behind, and
    Excel asks about links on every open because it cannot resolve them.
    The names are part of the removed sheet's machinery, so they go with
    it.

    Args:
        wb: The open workbook.
        removed: The names of the sheets that were deleted.
    """
    targets = {f"{name}!" for name in removed} | {f"'{name}'!" for name in removed}

    dangling = [
        key
        for key, defined in wb.defined_names.items()
        if any(target in str(defined.value) for target in targets)
    ]
    for key in dangling:
        del wb.defined_names[key]

    for ws in wb.worksheets:
        local = [
            key
            for key, defined in ws.defined_names.items()
            if any(target in str(defined.value) for target in targets)
        ]
        for key in local:
            del ws.defined_names[key]
        dangling.extend(local)

    if dangling:
        print(f"  Removed {len(dangling)} defined names that pointed at them.")


def clear_calculated_columns(wb: openpyxl.Workbook) -> None:
    """Forget the column formulas of tables whose cells now hold values.

    A table column can carry a formula that Excel expects every cell of
    that column to repeat. Once the formulas have been resolved to their
    results, the cells are constants and the table still asks for the
    formula, which Excel flags on every one of them as an inconsistent
    column. The formula is no longer true of the file, so it goes.

    Args:
        wb: The open workbook.
    """
    cleared = 0
    for ws in wb.worksheets:
        for table in ws.tables.values():
            for column in table.tableColumns:
                if column.calculatedColumnFormula is not None:
                    column.calculatedColumnFormula = None
                    cleared += 1
    if cleared:
        print(f"  Cleared {cleared} table column formulas that no longer describe the cells.")


def drop_hyperlink(cell) -> None:
    """Remove the link on a cell whose value has just been replaced.

    A hyperlink is not stored in the cell. It is a relationship of the
    sheet, and it survives every change to the value it sits under — so
    a column of email addresses replaced by keys keeps the addresses, in
    a part of the file no comparison of cell values will ever look at.
    The `loan_agent_contact` column of one workbook carried sixteen of
    them into the output, complete with the names in front of the `@`.

    Once the value is a key, whatever the link pointed at is not what the
    cell says any more, so it goes.

    Args:
        cell: The cell that was just written.
    """
    if cell.hyperlink is not None:
        cell.hyperlink = None


def report_surviving_hyperlinks(wb: openpyxl.Workbook) -> None:
    """Say where links are left, without saying what they point at.

    Anything still linked was not replaced, which may be right — a link
    to a public register is not client data. It has to be visible, so it
    is counted per sheet. The targets themselves are not printed: they
    are exactly the sort of thing that must not leave the machine.

    Args:
        wb: The open workbook.
    """
    remaining = {ws.title: len(ws._hyperlinks) for ws in wb.worksheets if ws._hyperlinks}
    if remaining:
        where = ", ".join(f"{name} ({count})" for name, count in sorted(remaining.items()))
        print(f"  WARNING: {sum(remaining.values())} hyperlinks are still in the file: {where}.")
        print("           A link keeps its target when the cell under it is replaced.")


def clear_document_properties(wb: openpyxl.Workbook) -> None:
    """Empty the properties of the file itself.

    A workbook records who created it and who saved it last, by full
    name, and it carries whatever custom properties the organisation's
    systems attached — in one file the classification label of a document
    protection service, with the email address of the person who applied
    it and the tenant it belongs to.

    None of that is in a cell, so nothing that compares cells will find
    it. Two derived files were already called clean while naming four
    people between them.

    The dates are kept: they say nothing about anyone, and a file with no
    timestamps looks stranger than it is.

    Args:
        wb: The open workbook.
    """
    properties = wb.properties
    named = [properties.creator, properties.lastModifiedBy, properties.lastPrinted]
    for field in ("creator", "lastModifiedBy", "title", "subject", "description", "keywords", "category", "identifier", "language", "lastPrinted", "revision", "version"):
        if hasattr(properties, field):
            setattr(properties, field, None)

    custom = len(wb.custom_doc_props.props) if wb.custom_doc_props is not None else 0
    if custom:
        wb.custom_doc_props.props = []

    if any(named) or custom:
        print(f"  Cleared the document properties, including {custom} custom ones.")


def strip_comments(wb: openpyxl.Workbook) -> int:
    """Remove every cell comment, and say how many there were.

    A comment is prose hanging off a cell, and prose cannot be checked
    for names cheaply — which is the same argument that puts free text
    columns into a replacement group rather than leaving them alone. In
    one workbook the comments named five companies of a borrower group,
    an arranging bank, a colleague by first name, and the date a payment
    arrived late.

    Nothing reads them. They are notes between the people who keep the
    file, so they are removed rather than replaced: a key in place of a
    sentence tells the reader of the derived file nothing either.

    Args:
        wb: The open workbook.

    Returns:
        The number of comments removed.
    """
    removed = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment is not None:
                    cell.comment = None
                    removed += 1
    if removed:
        print(f"  Removed {removed} cell comments; they are notes between people, and prose cannot be checked.")
    return removed


def demote_query_tables(wb: openpyxl.Workbook) -> None:
    """Turn tables that were fed by a query into ordinary tables.

    A table can be the result of an external query — a currency history
    pulled from a data source, say. The workbook then holds the query and
    the connection beside the table, and the table declares itself a
    query table that refers to them.

    openpyxl does not carry the query or the connection over, but it does
    write the declaration back. Excel opens the file, finds a table that
    says it is fed by a query that is not there, and offers to repair the
    workbook — which is a frightening thing to hand someone together with
    the assurance that the file is sound.

    The data is already in the sheet and nothing is going to refresh it
    here, so the honest form is a plain table.

    Args:
        wb: The open workbook.
    """
    demoted = 0
    for ws in wb.worksheets:
        for table in ws.tables.values():
            if table.tableType is None:
                continue
            table.tableType = None
            table.connectionId = None
            for column in table.tableColumns:
                column.queryTableFieldId = None
            demoted += 1
    if demoted:
        print(f"  Turned {demoted} query tables into plain ones; the query itself is not carried over.")


def drop_external_links(wb: openpyxl.Workbook) -> None:
    """Remove the links to other workbooks, and what they last read.

    A formula that reaches into another workbook does not only record
    where that workbook is — it caches the values it last read from it,
    so the figure still shows when the other file is not there. Both
    survive anonymisation untouched: the cells being replaced are in this
    workbook, and the cache is not a cell.

    So the copy carries a directory path, often with a person's name in
    it, and a block of someone else's data that nothing here has ever
    looked at. The PE tracker has thirty of them, one caching 792 values.

    Nothing is lost by removing them. The formulas are resolved to their
    results before this runs, and a derived dataset is not going to
    refresh anything.

    Args:
        wb: The open workbook.
    """
    count = len(wb._external_links)
    if count:
        wb._external_links = []
        print(f"  Removed {count} links to other workbooks, with the values they had cached.")


def drop_pivot_caches(wb: openpyxl.Workbook) -> None:
    """Remove pivot tables and the copy of the data they keep.

    A pivot table does not read the sheet when it draws; it reads a cache
    of the source range stored beside it. Anonymising the sheet leaves
    that cache holding the original rows.

    Args:
        wb: The open workbook.
    """
    count = len(wb._pivots)
    for ws in wb.worksheets:
        ws._pivots = []
    if count:
        wb._pivots = []
        print(f"  Removed {count} pivot tables and their cached source data.")


def drop_defined_names(wb: openpyxl.Workbook) -> None:
    """Remove the named ranges.

    A defined name is a label on a range, and the label is written by
    hand — often after whatever the range is about, which is a name of
    the kind this tool exists to remove. It stays that way whether or not
    the cells below it still say the same. Some of them point into other
    workbooks and carry that path with them.

    They are only removed where the formulas have been resolved to their
    results — nothing reads a name in a workbook that no longer computes
    anything. In ``keep`` mode they stay, because removing them would
    break every formula that uses one.

    Args:
        wb: The open workbook.
    """
    count = len(wb.defined_names)
    if count:
        for name in list(wb.defined_names):
            del wb.defined_names[name]
        print(f"  Removed {count} named ranges; their names are written by hand.")
    for ws in wb.worksheets:
        local = len(ws.defined_names)
        if local:
            for name in list(ws.defined_names):
                del ws.defined_names[name]
            print(f"  Removed {local} named ranges local to a sheet.")


def check_every_sheet_is_accounted_for(
    wb: openpyxl.Workbook, config: dict, mode: str, excel_path: Path
) -> None:
    """Report sheets that neither a group nor `ignore_sheets` mentions.

    A configuration is written against the file you had. The next export
    can carry a sheet nobody has seen, and it passes through untouched —
    with its values — because nothing refers to it. That is the failure
    this guards: not a wrong rule, but a missing one.

    A sheet counts as accounted for if a group names it, if
    ``ignore_sheets`` removes it, or if ``keep_sheets`` says it stays as
    it is. The third is needed as often as the other two: a sheet can be
    empty, or hold nothing but a list of currency codes.

    ``warn`` counts them and continues; ``fail`` refuses. Use ``fail``
    wherever the output is supposed to be free of original values, which
    can only be claimed for a file whose every sheet was considered.

    The names go into a file beside the workbook rather than onto the
    screen. In a workbook of this kind a sheet is often named after the
    transaction it holds, so the list of sheets nobody has decided about
    is itself confidential — and it is the one thing whoever runs this is
    most likely to copy somewhere in order to ask about it.

    Args:
        wb: The workbook, after the ignored sheets have been removed.
        config: The parsed configuration.
        mode: ``warn`` or ``fail``.
        excel_path: The input file; the listing is written beside it.

    Raises:
        ValueError: In ``fail`` mode, if any sheet is unaccounted for.
    """
    if mode not in ("warn", "fail"):
        raise ValueError(f"unlisted_sheets: {mode!r} — expected 'warn' or 'fail'")

    named: set[str] = set()
    for group in config.get("groups", []):
        for col in group.get("columns", []):
            named.update(matching_sheets(wb, col, warn_if_absent=False))
    for entry in config.get("keep_sheets", []):
        named.update(matching_sheets(wb, entry, warn_if_absent=False))

    unlisted = [name for name in wb.sheetnames if name not in named]
    if not unlisted:
        return

    listing = excel_path.with_name(f"{excel_path.stem}_unlisted_sheets.txt")
    listing.write_text("\n".join(unlisted) + "\n", encoding="utf-8")

    print(f"  {len(unlisted)} sheets are in the file but in no group and not in ignore_sheets.")
    print(f"  Their names are in {listing.name}, beside the workbook.")
    if mode == "fail":
        raise ValueError(
            f"{len(unlisted)} sheets are unaccounted for; list them in a group or in ignore_sheets. "
            f"See {listing.name}."
        )


def check_patterns_reach_one_shape(wb: openpyxl.Workbook, config: dict, mode: str, excel_path: Path) -> None:
    """Check that the sheets a pattern matches all have the same header row.

    A pattern plus a column letter is a bet: that every sheet the pattern
    reaches puts the same thing in the same place. Where the bet is wrong
    the letters land on the neighbouring column, and nothing says so —
    the count of replaced values goes up, the structural comparison finds
    the shape unchanged, and the values that were supposed to be replaced
    are still there.

    So the bet is checked, and only where it was made: the header of the
    first matching sheet is taken as the shape, and every other matching
    sheet is compared against it in exactly the columns some entry names
    by letter. Nowhere else. A column the configuration never touches may
    hold anything — one cashflow tab of TF1 has a leftover heading in A,
    which says nothing about the columns that are addressed — and the
    scratch area to the right of the data differs from tab to tab by
    design, which is why it is addressed as a range rather than by
    letter.

    What is reported is a count and a column number. The names of the
    sheets that differ go into a file beside the workbook, because a
    sheet is often named after the transaction it holds.

    Args:
        wb: The workbook, after the ignored sheets have been removed.
        config: The parsed configuration.
        mode: ``warn`` or ``fail``.
        excel_path: The input file; the listing is written beside it.

    Raises:
        ValueError: In ``fail`` mode, if a pattern reaches two shapes.
    """
    reach: dict[str, set[int]] = {}
    for group in config.get("groups", []):
        for col in group.get("columns", []):
            pattern = col.get("sheet_pattern")
            if pattern is None or "col" not in col:
                continue
            index = openpyxl.utils.column_index_from_string(col["col"])
            reach.setdefault(pattern, set()).add(index)

    odd: list[str] = []
    for pattern, indices in reach.items():
        names = matching_sheets(wb, {"sheet_pattern": pattern}, warn_if_absent=False)
        if len(names) < 2:
            continue
        columns = sorted(indices)
        shape = _header(wb[names[0]], columns)
        differing = [name for name in names[1:] if _header(wb[name], columns) != shape]
        if differing:
            elsewhere = _header(wb[differing[0]], columns)
            moved = ", ".join(
                openpyxl.utils.get_column_letter(index)
                for index, (here, there) in zip(columns, zip(shape, elsewhere))
                if here != there
            )
            print(
                f"  {len(differing)} of {len(names)} sheets matched by a pattern "
                f"do not have the header the letters assume, in column {moved}."
            )
            odd.extend(differing)

    if not odd:
        return

    listing = excel_path.with_name(f"{excel_path.stem}_differing_sheets.txt")
    listing.write_text("\n".join(odd) + "\n", encoding="utf-8")
    print(f"  Their names are in {listing.name}, beside the workbook.")
    if mode == "fail":
        raise ValueError(
            f"{len(odd)} sheets do not have the layout their pattern assumes; "
            f"address them separately. See {listing.name}."
        )


def _header(ws, columns: list[int]) -> tuple:
    """Return the header cells at the given one-based columns, lowercased."""
    row = next(ws.iter_rows(min_row=1, max_row=1, max_col=max(columns), values_only=True), ())
    return tuple(
        str(row[index - 1]).strip().lower() if index <= len(row) and row[index - 1] is not None else ""
        for index in columns
    )


def scale_value(value: object, factor: float) -> object | None:
    """Return the value multiplied by the factor, keeping the type it had.

    An integer stays an integer: writing 1283.7 where the source held
    1000 changes the shape of the column, not just its content, and any
    check comparing the two files would report it.

    Booleans are numbers in Python but flags in a spreadsheet, so they
    are left alone.

    Args:
        value: The original cell value.
        factor: The factor to multiply by.

    Returns:
        The scaled value, or None if the cell holds nothing numeric.
    """
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    scaled = value * factor
    return int(round(scaled)) if isinstance(value, int) else round(scaled, 6)


def iter_cells(wb: openpyxl.Workbook, columns: list[dict], group_name: str = "") -> list[tuple]:
    """Yield every cell of the configured columns as (worksheet, row, column).

    A column entry that matches no sheet stops the run. It is the one
    mistake this tool cannot survive quietly: the values it was meant to
    replace stay in the file, and every other check still passes. The
    configuration says the column is handled, the structural comparison
    finds no difference because nothing moved, and the residual check
    only looks at text — so a mistyped pattern over a sheet full of
    amounts goes through unseen. It did, once.

    Args:
        wb: The open workbook.
        columns: The ``columns`` entries of one group.
        group_name: The group these columns belong to, for the error.

    Returns:
        A list of (worksheet, row index, column index) tuples.

    Raises:
        ValueError: If a column entry matches no sheet in the workbook.
    """
    cells: list[tuple] = []
    for col_spec in columns:
        data_from = col_spec.get("data_from_row", 2)
        described = col_spec.get("col") or f"{col_spec.get('col_from')} onwards"

        sheets = matching_sheets(wb, col_spec, warn_if_absent=False)
        if not sheets:
            where = col_spec.get("sheet") or col_spec.get("sheet_pattern") or "empty: true"
            raise ValueError(
                f"group '{group_name}': column {described} refers to {where!r}, "
                f"which matches no sheet in the file. Nothing would be replaced there."
            )

        for sheet_name in sheets:
            ws = wb[sheet_name]
            for col_idx in column_range(col_spec, ws):
                for row_idx in range(data_from, ws.max_row + 1):
                    cells.append((ws, row_idx, col_idx))
    return cells


def column_range(col_spec: dict, ws) -> range:
    """Return the column indices one entry covers in one sheet.

    An entry names a single column (``col``) or everything from one
    column rightwards (``col_from``, optionally bounded by ``col_to``).

    The range is what makes the working area of a sheet addressable. The
    cashflow tabs of the DataSet workbooks hold their data in A to AB and
    everything to the right of that is scratch: running sums, repeated
    drawdowns, a second copy of the cashflow. Those columns carry the
    same amounts as the data columns, so leaving them alone leaves the
    originals in the file — and they cannot be listed by letter, because
    they sit at different letters in each tab. What holds across all of
    them is where the data stops.

    Args:
        col_spec: One ``columns`` entry.
        ws: The worksheet the range is resolved against.

    Returns:
        The column indices, one-based.
    """
    if "col" in col_spec:
        index = openpyxl.utils.column_index_from_string(col_spec["col"])
        return range(index, index + 1)

    first = openpyxl.utils.column_index_from_string(col_spec["col_from"])
    last = ws.max_column
    if "col_to" in col_spec:
        last = min(last, openpyxl.utils.column_index_from_string(col_spec["col_to"]))
    return range(first, last + 1)


def apply_key_group(wb: openpyxl.Workbook, group: dict) -> dict[str, str]:
    """Replace the values of one group with stable keys.

    Args:
        wb: The open workbook.
        group: One entry of the ``groups`` list.

    Returns:
        The mapping from original value to key.
    """
    if "prefix" not in group:
        raise ValueError(f"group {group['name']!r}: a key group needs a prefix")

    prefix = group["prefix"]
    numbers_too = group.get("include_numbers", False)
    dates_too = group.get("include_dates", False)
    mapping: dict[str, str] = OrderedDict()
    replacements: list[tuple] = []

    for ws, row_idx, col_idx in iter_cells(wb, group["columns"], group["name"]):
        cell = ws.cell(row=row_idx, column=col_idx)
        if isinstance(cell.value, (int, float)) and not numbers_too:
            # A key is text. Writing one over a number changes the type of
            # the column, which is a change to the file's shape rather than
            # to its content — and columns holding text among numbers are
            # common enough that keying them silently would be a trap. A
            # group that does mean to replace numbers says `include_numbers`.
            continue
        if isinstance(cell.value, (datetime.datetime, datetime.date, datetime.time)) and not dates_too:
            # The same argument, and one more: a group that covers a range
            # of columns will meet dates whether it meant to or not. Where
            # a date does have to go, `include_dates` says so — but it
            # leaves a date column full of text behind, so shifting the
            # dates is usually the better answer.
            continue
        val = str(cell.value).strip() if cell.value is not None else ""
        if not val or val in ("None", "nan"):
            continue
        if val not in mapping:
            mapping[val] = f"{prefix}{len(mapping) + 1:04d}"
        replacements.append((ws, row_idx, col_idx, mapping[val]))

    for ws, row_idx, col_idx, new_val in replacements:
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = new_val
        drop_hyperlink(cell)

    print(f"  Group '{group['name']}': {len(mapping)} unique values replaced.")
    return mapping


def apply_scale_group(wb: openpyxl.Workbook, group: dict) -> None:
    """Multiply the numbers of one group by the group's factor.

    One factor for a whole workbook keeps the relations between amounts
    intact — a commitment stays below its tranche total, and a sum of
    cashflows still reconciles. Ratios must therefore not be scaled: a
    set of weights adding up to 1.0 would add up to the factor. Which
    columns are amounts and which are ratios is a decision about the
    data, which is why it is made in the configuration.

    Scaling is not reversible: rounding to the source type loses the
    remainder. Scale groups are therefore not written to the mapping
    file.

    Args:
        wb: The open workbook.
        group: One entry of the ``groups`` list.
    """
    factor = group.get("factor")
    if factor is None:
        raise ValueError(f"group {group['name']!r}: a scale group needs a factor")

    replacements: list[tuple] = []
    skipped = 0

    for ws, row_idx, col_idx in iter_cells(wb, group["columns"], group["name"]):
        value = ws.cell(row=row_idx, column=col_idx).value
        if value is None or value == "":
            continue
        scaled = scale_value(value, factor)
        if scaled is None:
            skipped += 1
            continue
        replacements.append((ws, row_idx, col_idx, scaled))

    for ws, row_idx, col_idx, new_val in replacements:
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = new_val
        drop_hyperlink(cell)

    note = f", {skipped} non-numeric cells left alone" if skipped else ""
    print(f"  Group '{group['name']}': {len(replacements)} numbers scaled by {factor}{note}.")


def output_path(excel_path: Path, suffix: str) -> Path:
    """Where the anonymised copy is written.

    A macro-enabled workbook is never written back as one. openpyxl does
    not carry the macro project over, so an ``.xlsm`` output would
    declare itself macro-enabled and hold no macros — the file Excel
    offers to repair. The derived dataset does not need them either: what
    is taken over from it is data, not automation.

    Args:
        excel_path: Path to the input Excel file (``.xlsx`` or ``.xlsm``).
        suffix: What to append to the stem, from ``output_suffix``.

    Returns:
        The output path, always with an ``.xlsx`` extension.
    """
    out_path = excel_path.with_stem(excel_path.stem + suffix)
    if out_path.suffix.lower() == ".xlsm":
        out_path = out_path.with_suffix(".xlsx")
    return out_path


def anonymize(excel_path: Path, config_path: Path) -> None:
    """Replace configured cell values with stable keys and optionally save the mapping.

    Args:
        excel_path: Path to the input Excel file (``.xlsx`` or ``.xlsm``).
        config_path: Path to the YAML config defining groups, prefixes, and columns.
    """
    config = yaml.safe_load(config_path.read_text(encoding="utf-8"))

    suffix = config.get("output_suffix", "_anonymized")
    map_file = config.get("save_mapping")
    formulas = config.get("formulas", "values")
    out_path = output_path(excel_path, suffix)
    if excel_path.suffix.lower() == ".xlsm":
        print(f"  Macro-enabled input: the copy is written as {out_path.name}, without the macros.")

    wb = load_source(excel_path, formulas)

    if formulas == "values":
        uncached = count_uncached_formulas(excel_path)
        if uncached:
            print(f"  WARNING: {uncached} formula cells carry no calculated result and arrive empty.")
            print("           Open the source in Excel, let it recalculate, save, and run again.")

    drop_sheets(wb, config.get("ignore_sheets", []))
    check_every_sheet_is_accounted_for(wb, config, config.get("unlisted_sheets", "warn"), excel_path)
    check_patterns_reach_one_shape(wb, config, config.get("unlisted_sheets", "warn"), excel_path)

    demote_query_tables(wb)
    drop_external_links(wb)
    drop_pivot_caches(wb)
    clear_document_properties(wb)

    if config.get("comments", "drop") == "drop":
        strip_comments(wb)
    else:
        print("  WARNING: cell comments are kept. Nothing checks what is written in them.")

    if formulas == "values":
        clear_calculated_columns(wb)
        drop_defined_names(wb)

    full_mapping: dict[str, dict[str, str]] = {}

    for group in config.get("groups", []):
        strategy = group.get("strategy", "key")
        if strategy == "key":
            full_mapping[group["name"]] = apply_key_group(wb, group)
        elif strategy == "scale":
            apply_scale_group(wb, group)
        else:
            raise ValueError(f"group {group['name']!r}: unknown strategy {strategy!r}")

    report_surviving_hyperlinks(wb)

    wb.save(out_path)
    print(f"Anonymized file : {out_path}")

    if map_file:
        map_path = excel_path.parent / map_file
        # Mapping contains original values – keep local, never commit
        map_path.write_text(
            json.dumps(full_mapping, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"Mapping saved   : {map_path}")


def main() -> None:
    """Parse CLI arguments and run the anonymization."""
    parser = argparse.ArgumentParser(
        description="Anonymize Excel files by replacing cell values with unique keys"
    )
    parser.add_argument("excel", type=Path, help="Input Excel file (.xlsx or .xlsm)")
    _base = Path(sys.executable).parent if getattr(sys, "frozen", False) else Path(__file__).parent.parent
    parser.add_argument(
        "--config", type=Path,
        default=_base / "anonymize_config.yaml",
        help="Config file (default: anonymize_config.yaml in project root)",
    )
    args = parser.parse_args()

    if not args.excel.exists():
        print(f"ERROR: file not found: {args.excel}")
        raise SystemExit(1)
    if args.excel.suffix.lower() not in (".xlsx", ".xlsm"):
        print(f"ERROR: only .xlsx and .xlsm are supported: {args.excel}")
        raise SystemExit(1)
    if not args.config.exists():
        print(f"ERROR: config not found: {args.config}")
        raise SystemExit(1)

    print(f"Input : {args.excel}")
    print(f"Config: {args.config}")
    anonymize(args.excel, args.config)


if __name__ == "__main__":
    main()
