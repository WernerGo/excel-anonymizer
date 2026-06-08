"""
Generate a sample Excel file from an anonymize config.

Creates realistic fake data for all configured sheets and columns.
Values are drawn from a small pool (with repetition) so the same value
appears across multiple columns — exactly the relationship-preservation
scenario the anonymizer is designed for.

Usage:
    python src/generate_sample.py
    python src/generate_sample.py --config examples/config_names.yaml
    python src/generate_sample.py --config examples/config_full.yaml --rows 50 --output my_sample.xlsx
"""

import argparse
import random
import sys
from pathlib import Path

import openpyxl
import yaml
from faker import Faker

from faker_utils import infer_faker_type, make_generator


def generate(config_path: Path, output_path: Path, rows: int = 30) -> None:
    config = yaml.safe_load(config_path.read_text(encoding="utf-8"))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Collect sheets and their column/data-row info
    sheet_cols: dict[str, dict[str, int]] = {}  # sheet → {col_letter → data_from_row}
    for group in config.get("groups", []):
        for col_spec in group["columns"]:
            sheet = col_spec["sheet"]
            col   = col_spec["col"]
            dfr   = col_spec.get("data_from_row", 2)
            sheet_cols.setdefault(sheet, {})[col] = dfr

    # Create worksheets
    worksheets: dict[str, openpyxl.worksheet.worksheet.Worksheet] = {
        name: wb.create_sheet(name) for name in sheet_cols
    }

    locale = config.get("locale", "en_US")
    fake   = Faker(locale)

    # Fill columns group by group
    for group in config.get("groups", []):
        group_name  = group["name"]
        faker_type  = group.get("faker_type") or infer_faker_type(group_name)
        generate    = make_generator(faker_type, fake)

        # Small pool → many repetitions → relationships to preserve
        pool_size = max(4, rows // 5)
        pool = [generate() for _ in range(pool_size)]

        for col_spec in group["columns"]:
            sheet_name = col_spec["sheet"]
            col_letter = col_spec["col"]
            data_from  = col_spec.get("data_from_row", 2)

            if sheet_name not in worksheets:
                continue

            ws      = worksheets[sheet_name]
            col_idx = openpyxl.utils.column_index_from_string(col_letter)

            # Header row(s): write column label one row above data
            ws.cell(row=data_from - 1, column=col_idx, value=f"{group_name} [{col_letter}]")

            # Data rows
            for offset in range(rows):
                ws.cell(row=data_from + offset, column=col_idx, value=random.choice(pool))

    wb.save(output_path)
    print(f"Sample file written : {output_path}  ({rows} rows per sheet)")
    print(f"Test anonymizer     : python src/anonymize.py {output_path} --config {config_path}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate a sample Excel file from an anonymize config"
    )
    _base = Path(sys.executable).parent if getattr(sys, "frozen", False) else Path(__file__).parent.parent
    parser.add_argument(
        "--config", type=Path,
        default=_base / "examples" / "config_names.yaml",
        help="Anonymize config to base the sample on (default: examples/config_names.yaml)",
    )
    parser.add_argument(
        "--output", type=Path,
        default=_base / "sample_data.xlsx",
        help="Output file (default: sample_data.xlsx in project root)",
    )
    parser.add_argument(
        "--rows", type=int, default=30,
        help="Number of data rows per sheet (default: 30)",
    )
    args = parser.parse_args()

    if not args.config.exists():
        print(f"ERROR: config not found: {args.config}")
        raise SystemExit(1)

    generate(args.config, args.output, args.rows)


if __name__ == "__main__":
    main()
