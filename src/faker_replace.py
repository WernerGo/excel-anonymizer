"""
Excel anonymization with realistic fake data.

Like anonymize.py but replaces values with realistic fake data (via Faker)
instead of abstract keys. Identical values within a group always get the
same fake replacement, preserving relationships across columns
(e.g. employee → manager links stay intact).

Requires a `faker_type` per group in the config. Supported types:
  last_name | first_name | full_name | company | city | department | email | word

Usage:
    uv run src/faker_replace.py input.xlsx
    uv run src/faker_replace.py input.xlsx --config examples/config_faker.yaml
    uv run src/faker_replace.py input.xlsx --locale de_DE
"""

import argparse
import json
import sys
from collections import OrderedDict
from pathlib import Path

import openpyxl
import yaml
from faker import Faker

from faker_utils import make_generator


def faker_replace(excel_path: Path, config_path: Path, locale_override: str | None = None) -> None:
    config = yaml.safe_load(config_path.read_text(encoding="utf-8"))

    locale = locale_override or config.get("locale", "en_US")
    fake   = Faker(locale)
    Faker.seed(0)  # reproducible output for the same input file

    suffix   = config.get("output_suffix", "_faker")
    map_file = config.get("save_mapping")
    out_path = excel_path.with_stem(excel_path.stem + suffix)

    wb = openpyxl.load_workbook(excel_path)
    full_mapping: dict[str, dict[str, str]] = {}

    for group in config.get("groups", []):
        group_name = group["name"]
        faker_type = group.get("faker_type", "word")

        try:
            generate = make_generator(faker_type, fake)
        except ValueError as e:
            print(f"  ERROR in group '{group_name}': {e}")
            continue

        mapping: dict[str, str] = OrderedDict()
        cell_refs: list[tuple] = []

        for col_spec in group["columns"]:
            sheet_name = col_spec["sheet"]
            col_letter = col_spec["col"]
            data_from  = col_spec.get("data_from_row", 2)

            if sheet_name not in wb.sheetnames:
                print(f"  WARNING: sheet '{sheet_name}' not found, skipped.")
                continue

            ws      = wb[sheet_name]
            col_idx = openpyxl.utils.column_index_from_string(col_letter)

            for row_idx in range(data_from, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val  = str(cell.value).strip() if cell.value is not None else ""
                if not val or val in ("None", "nan"):
                    continue
                if val not in mapping:
                    mapping[val] = generate()
                cell_refs.append((ws, row_idx, col_idx, mapping[val]))

        for ws, row_idx, col_idx, new_val in cell_refs:
            ws.cell(row=row_idx, column=col_idx).value = new_val

        full_mapping[group_name] = mapping
        print(f"  Group '{group_name}' [{faker_type}]: {len(mapping)} unique values replaced.")

    wb.save(out_path)
    print(f"Anonymized file : {out_path}")

    if map_file:
        map_path = excel_path.parent / map_file
        map_path.write_text(
            json.dumps(full_mapping, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"Mapping saved   : {map_path}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Anonymize Excel files with realistic fake data via Faker"
    )
    _base = Path(sys.executable).parent if getattr(sys, "frozen", False) else Path(__file__).parent.parent
    parser.add_argument("excel", type=Path, help="Input Excel file (.xlsx)")
    parser.add_argument(
        "--config", type=Path,
        default=_base / "examples" / "config_faker.yaml",
        help="Config file (default: examples/config_faker.yaml)",
    )
    parser.add_argument(
        "--locale", type=str, default=None,
        help="Faker locale, e.g. de_DE, en_US, fr_FR (overrides config)",
    )
    args = parser.parse_args()

    if not args.excel.exists():
        print(f"ERROR: file not found: {args.excel}")
        raise SystemExit(1)
    if not args.config.exists():
        print(f"ERROR: config not found: {args.config}")
        raise SystemExit(1)

    print(f"Input : {args.excel}")
    print(f"Config: {args.config}")
    print(f"Locale: {args.locale or '(from config)'}")
    faker_replace(args.excel, args.config, args.locale)


if __name__ == "__main__":
    main()
